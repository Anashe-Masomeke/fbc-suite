"""
FBC Suite — Combined Desktop App
──────────────────────────────────
  📊  Sharestock Upload Converter   (Tab 1)
  ✉   Deal Note Email Automator    (Tab 2)

Requirements:
    pip install pandas openpyxl fpdf2 pywin32 pymupdf gspread google-auth

    updating new .exe file:
    pyinstaller --onefile --noconsole fbc_suite.py --name fbc-suite
"""

# ════════════════════════════════════════════════════════════════════════════
#  AUTO-UPDATE
# ════════════════════════════════════════════════════════════════════════════
import sys, os, subprocess, urllib.request

VERSION       = 27
GITHUB_USER   = "Anashe-Masomeke"
GITHUB_REPO   = "fbc-suite"
GITHUB_BRANCH = "main"
EXE_NAME      = "fbc-suite.exe"

_EXE = f"https://github.com/{GITHUB_USER}/{GITHUB_REPO}/releases/latest/download/{EXE_NAME}"
_VER = (f"https://raw.githubusercontent.com/"
        f"{GITHUB_USER}/{GITHUB_REPO}/{GITHUB_BRANCH}/version.txt")

def _remote_ver():
    try:
        with urllib.request.urlopen(_VER, timeout=4) as r:
            return int(r.read().decode().strip())
    except Exception:
        return -1

def check_and_apply_update():
    rv = _remote_ver()
    if rv <= VERSION:
        return

    current_exe = os.path.abspath(sys.argv[0])
    exe_dir     = os.path.dirname(current_exe)

    new_exe_path = os.path.join(exe_dir, f"fbc-suite-v{rv}.exe")
    bat_path     = os.path.join(exe_dir, "_fbc_updater.bat")

    import tkinter as tk
    from tkinter import messagebox

    # Tiny non-blocking splash so the app doesn't look frozen during download.
    splash = tk.Tk()
    splash.title("FBC Suite")
    splash.resizable(False, False)
    splash.configure(bg=SIDEBAR_BG)
    w, h = 320, 110
    x = (splash.winfo_screenwidth()  - w) // 2
    y = (splash.winfo_screenheight() - h) // 2
    splash.geometry(f"{w}x{h}+{x}+{y}")
    tk.Label(splash, text="Updating FBC Suite…", bg=SIDEBAR_BG, fg=WHITE,
             font=("Segoe UI", 11, "bold")).pack(pady=(22, 6))
    tk.Label(splash, text=f"v{VERSION}  →  v{rv}", bg=SIDEBAR_BG, fg="#90CAF9",
             font=("Segoe UI", 9)).pack()
    splash.update()

    try:
        MIN_SIZE = 20 * 1024 * 1024
        with urllib.request.urlopen(_EXE, timeout=180) as resp:
            with open(new_exe_path, "wb") as f:
                while True:
                    chunk = resp.read(65536)
                    if not chunk:
                        break
                    f.write(chunk)
                    splash.update()

        size = os.path.getsize(new_exe_path)
        if size < MIN_SIZE:
            os.remove(new_exe_path)
            raise Exception(f"Download incomplete ({size // 1024} KB).")

        bat_lines = [
            "@echo off",
            "ping 127.0.0.1 -n 4 > nul",
            f'start "" "{new_exe_path}"',
            "ping 127.0.0.1 -n 2 > nul",
            'del "%~f0"',
        ]
        with open(bat_path, "w") as f:
            f.write("\n".join(bat_lines) + "\n")

        subprocess.Popen(
            ["cmd.exe", "/c", bat_path],
            creationflags=subprocess.CREATE_NO_WINDOW,
            close_fds=True
        )
        splash.destroy()
        sys.exit(0)

    except Exception as e:
        splash.destroy()
        for fp in [new_exe_path, bat_path]:
            try: os.remove(fp)
            except Exception: pass
        # Silent failure — don't block the user, just continue on the old version.
        # Update will be retried automatically next time the app opens.
        print(f"[Auto-update] Failed silently, continuing on v{VERSION}: {e}")

# ════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ════════════════════════════════════════════════════════════════════════════
import re, json, csv, threading
import shutil as _shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

# ── Voice imports (graceful layered fallback) ────────────────────────────────
_VOICE_READY   = False
_sr            = None
_tts           = None
_whisper_model = None
_vosk_model    = None
_RECOGNISER    = "none"
_fuzz          = None

def _init_voice():
    global _VOICE_READY, _sr, _tts, _fuzz, _RECOGNISER
    try:
        import pyttsx3
        _tts = pyttsx3.init()
        _tts.setProperty("rate", 165)
    except Exception:
        pass
    try:
        import speech_recognition as sr
        _sr = sr
    except ImportError:
        return
    try:
        import whisper as _w
        _RECOGNISER = "whisper"
    except ImportError:
        pass
    if _RECOGNISER == "none":
        try:
            import vosk as _v  # noqa: F401
            _RECOGNISER = "vosk"
        except ImportError:
            pass
    if _RECOGNISER == "none":
        _RECOGNISER = "google"
    try:
        from rapidfuzz import process as rfp, fuzz as rff
        _fuzz = (rfp, rff)
    except ImportError:
        pass
    _VOICE_READY = True

threading.Thread(target=_init_voice, daemon=True).start()

_whisper_lock = threading.Lock()

def _get_whisper():
    global _whisper_model
    with _whisper_lock:
        if _whisper_model is None:
            import whisper
            _whisper_model = whisper.load_model("base.en")
    return _whisper_model

_vosk_lock = threading.Lock()
_VOSK_MODEL_PATH = os.path.join(os.path.expanduser("~"), "vosk-model-small-en-us-0.15")

def _get_vosk():
    global _vosk_model
    with _vosk_lock:
        if _vosk_model is None:
            import vosk
            if not os.path.isdir(_VOSK_MODEL_PATH):
                raise FileNotFoundError(
                    f"Vosk model not found at:\n{_VOSK_MODEL_PATH}\n\n"
                    "Download from: alphacephei.com/vosk/models\n"
                    "Extract to your home folder.")
            _vosk_model = vosk.Model(_VOSK_MODEL_PATH)
    return _vosk_model

def _require(pkg, install_name=None):
    import importlib
    try:
        return importlib.import_module(pkg)
    except ImportError:
        name = install_name or pkg
        raise ImportError(
            f"Missing package '{name}'. Open terminal and run:\n  pip install {name}")

# ════════════════════════════════════════════════════════════════════════════
#  SHARED COLOURS
# ════════════════════════════════════════════════════════════════════════════
FBC_DARK   = "#003B6F"
FBC_MID    = "#0066B3"
FBC_ACCENT = "#00A3E0"
GREEN_DARK = "#1A6B3A"
RED_DARK   = "#B71C1C"
WHITE      = "#FFFFFF"
BG         = "#F0F4F8"
CARD_BG    = "#FFFFFF"
SEP_CLR    = "#D0DAE8"
TAG_BLUE   = "#E8F1FB"
COL1_HDR   = "#003B6F"
COL2_HDR   = "#1A3A6B"
BOTTOM     = "#0D2B4E"
SIDEBAR_BG      = "#001F3F"
SIDEBAR_ACTIVE  = "#0066B3"
SIDEBAR_HOVER   = "#003B6F"
SIDEBAR_TEXT    = "#B0C8E8"
SIDEBAR_TEXT_ON = "#FFFFFF"

# ════════════════════════════════════════════════════════════════════════════
#  LOGIN DIALOG
# ════════════════════════════════════════════════════════════════════════════
APP_PASSWORD = "enock"
MAX_ATTEMPTS = 6

class LoginDialog(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("FBC Suite — Login")
        self.resizable(False, False)
        self.configure(bg=SIDEBAR_BG)
        self._attempts = 0
        self.authenticated = False
        self._build()
        self.update_idletasks()
        w, h = 380, 340
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build(self):
        hdr = tk.Frame(self, bg=FBC_ACCENT, pady=18)
        hdr.pack(fill="x")
        tk.Label(hdr, text="FBC", bg=FBC_DARK, fg=WHITE,
                 font=("Segoe UI", 20, "bold"), padx=12, pady=6).pack()
        tk.Label(hdr, text="Suite", bg=FBC_ACCENT, fg=WHITE,
                 font=("Segoe UI", 11)).pack(pady=(2, 0))
        body = tk.Frame(self, bg=SIDEBAR_BG, padx=36, pady=24)
        body.pack(fill="both", expand=True)
        tk.Label(body, text="Enter Password", bg=SIDEBAR_BG, fg=SIDEBAR_TEXT,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")
        pw_row = tk.Frame(body, bg=SIDEBAR_BG)
        pw_row.pack(fill="x", pady=(6, 0))
        self._pw_var = tk.StringVar()
        self._show_pw = False
        self.entry_pw = tk.Entry(pw_row, textvariable=self._pw_var, show="●",
                                 font=("Segoe UI", 12), bg="#0D2B4E", fg=WHITE,
                                 insertbackground=WHITE, relief="flat",
                                 highlightbackground=FBC_MID, highlightthickness=1)
        self.entry_pw.pack(side="left", fill="x", expand=True, ipady=8, padx=(0, 4))
        self.entry_pw.focus()
        self.btn_eye = tk.Button(pw_row, text="👁", command=self._toggle_show,
                                 bg="#0D2B4E", fg=SIDEBAR_TEXT, relief="flat",
                                 font=("Segoe UI", 12), cursor="hand2",
                                 activebackground=FBC_MID, activeforeground=WHITE,
                                 padx=6)
        self.btn_eye.pack(side="left")
        self.lbl_err = tk.Label(body, text="", bg=SIDEBAR_BG, fg="#FF6B6B",
                                font=("Segoe UI", 9))
        self.lbl_err.pack(anchor="w", pady=(6, 0))
        self.lbl_attempts = tk.Label(body, text="", bg=SIDEBAR_BG, fg="#607080",
                                     font=("Segoe UI", 8))
        self.lbl_attempts.pack(anchor="w")
        self.btn_login = tk.Button(body, text="  🔓  Login  ",
                                   command=self._attempt_login,
                                   bg=FBC_MID, fg=WHITE, relief="flat",
                                   font=("Segoe UI", 11, "bold"),
                                   cursor="hand2", pady=10, activebackground=FBC_ACCENT)
        self.btn_login.pack(fill="x", pady=(16, 0))
        self.entry_pw.bind("<Return>", lambda _: self._attempt_login())
        tk.Label(self, text=f"v{VERSION}", bg=SIDEBAR_BG, fg="#2A4A6A",
                 font=("Segoe UI", 8)).pack(side="bottom", pady=6)

    def _toggle_show(self):
        self._show_pw = not self._show_pw
        self.entry_pw.config(show="" if self._show_pw else "●")
        self.btn_eye.config(text="🙈" if self._show_pw else "👁")

    def _attempt_login(self):
        entered = self._pw_var.get().strip().lower()
        if entered == APP_PASSWORD.lower():
            self.authenticated = True
            self.destroy()
            return
        self._attempts += 1
        remaining = MAX_ATTEMPTS - self._attempts
        if remaining <= 0:
            messagebox.showerror("Access Denied",
                "Too many incorrect attempts.\nThe application will now close.")
            self.destroy()
            return
        self.lbl_err.config(text="❌  Incorrect password. Please try again.")
        self.lbl_attempts.config(
            text=f"  {remaining} attempt{'s' if remaining > 1 else ''} remaining")
        self._pw_var.set("")
        self.entry_pw.focus()
        self._shake()

    def _shake(self, times=6, distance=8):
        x0 = self.winfo_x()
        y0 = self.winfo_y()
        def step(n):
            if n == 0:
                self.geometry(f"+{x0}+{y0}")
                return
            offset = distance if n % 2 == 0 else -distance
            self.geometry(f"+{x0 + offset}+{y0}")
            self.after(40, lambda: step(n - 1))
        step(times)

    def _on_close(self):
        self.authenticated = False
        self.destroy()


# ════════════════════════════════════════════════════════════════════════════
#  ── RECIPIENTS CONFIG  (persistent per-user JSON files) ────────────────────
# ════════════════════════════════════════════════════════════════════════════

# --- Sarestock / Deals Confirmation recipients --------------------------------
SARESTOCK_RECIP_FILE = os.path.join(os.path.expanduser("~"), ".fbc_sarestock_recipients.json")

_SARESTOCK_DEFAULT_TO = ["Anesu.Zingundu@fbc.co.zw"]
_SARESTOCK_DEFAULT_CC = [
    "Enock.Rukarwa@fbc.co.zw", "Manatsa.Tagwireyi@fbc.co.zw",
    "Norman.Chirima@fbc.co.zw", "Richard.Mashava@fbc.co.zw",
    "Anashe.Masomeke@fbc.co.zw",
]

def load_sarestock_recipients():
    try:
        with open(SARESTOCK_RECIP_FILE) as f:
            d = json.load(f)
        return d.get("to", _SARESTOCK_DEFAULT_TO), d.get("cc", _SARESTOCK_DEFAULT_CC)
    except Exception:
        return list(_SARESTOCK_DEFAULT_TO), list(_SARESTOCK_DEFAULT_CC)

def save_sarestock_recipients(to_list, cc_list):
    with open(SARESTOCK_RECIP_FILE, "w") as f:
        json.dump({"to": to_list, "cc": cc_list}, f, indent=2)

# --- Custodian recipients overrides ------------------------------------------
CUSTODIAN_RECIP_FILE = os.path.join(os.path.expanduser("~"), ".fbc_custodian_recipients.json")

def load_custodian_overrides():
    """Returns dict: {custodian_code: {"to": [...], "cc": [...]}}"""
    try:
        with open(CUSTODIAN_RECIP_FILE) as f:
            return json.load(f)
    except Exception:
        return {}

def save_custodian_overrides(overrides):
    with open(CUSTODIAN_RECIP_FILE, "w") as f:
        json.dump(overrides, f, indent=2)

# --- Client email CC override ------------------------------------------------
CLIENT_CC_FILE = os.path.join(os.path.expanduser("~"), ".fbc_client_cc.json")

_FBC_CC_DEFAULT = [
    "Manatsa Tagwireyi <Manatsa.Tagwireyi@fbc.co.zw>",
    "Norman Chirima <Norman.Chirima@fbc.co.zw>",
    "Enock Rukarwa <Enock.Rukarwa@fbc.co.zw>",
    "Richard Mashava <Richard.Mashava@fbc.co.zw>",
    "Anesu Zingundu <Anesu.Zingundu@fbc.co.zw>",
]

def load_client_cc():
    try:
        with open(CLIENT_CC_FILE) as f:
            d = json.load(f)
        return d.get("cc", list(_FBC_CC_DEFAULT))
    except Exception:
        return list(_FBC_CC_DEFAULT)

def save_client_cc(cc_list):
    with open(CLIENT_CC_FILE, "w") as f:
        json.dump({"cc": cc_list}, f, indent=2)


# ════════════════════════════════════════════════════════════════════════════
#  SHARED RECIPIENTS EDITOR DIALOG
#  Reusable dialog for editing a To list + CC list (or just CC list).
#  Pass to_list=None to hide the To section (CC-only mode).
# ════════════════════════════════════════════════════════════════════════════
class RecipientsDialog(tk.Toplevel):
    """
    Generic editor for To / CC address lists.
    on_save(to_list, cc_list) is called when the user clicks Save.
    If to_list is None on open, the To section is hidden (CC-only).
    """
    def __init__(self, parent, title, to_list, cc_list, on_save,
                 to_label="To (primary recipients)",
                 cc_label="CC (copied recipients)"):
        super().__init__(parent)
        self.title(title)
        self.configure(bg=BG)
        self.resizable(False, False)
        self.grab_set()
        self._on_save = on_save
        self._has_to  = to_list is not None
        self._to_list = list(to_list) if to_list else []
        self._cc_list = list(cc_list)
        self._to_label = to_label
        self._cc_label = cc_label
        self._build()
        self.update_idletasks()
        w = 560
        h = self.winfo_reqheight()
        x = parent.winfo_rootx() + (parent.winfo_width()  - w) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build(self):
        tk.Label(self, text=f"  {self.title()}", bg=FBC_DARK, fg=WHITE,
                 font=("Segoe UI", 11, "bold"), pady=10).pack(fill="x")

        body = tk.Frame(self, bg=BG, padx=18, pady=12)
        body.pack(fill="both", expand=True)

        if self._has_to:
            self._to_frame = self._section(body, self._to_label, self._to_list)
        self._cc_frame = self._section(body, self._cc_label, self._cc_list)

        hint = tk.Label(body,
            text="One address per line.  Accepted formats:\n"
                 "  plain@email.com   or   Display Name <plain@email.com>",
            bg=BG, fg="#607080", font=("Segoe UI", 8), justify="left")
        hint.pack(anchor="w", pady=(6, 0))

        bot = tk.Frame(self, bg=BG, padx=18, pady=10)
        bot.pack(fill="x")
        tk.Button(bot, text="💾  Save", command=self._save,
                  bg=GREEN_DARK, fg=WHITE, relief="flat",
                  font=("Segoe UI", 10, "bold"), cursor="hand2",
                  padx=16, pady=7).pack(side="right")
        tk.Button(bot, text="Cancel", command=self.destroy,
                  bg="#607080", fg=WHITE, relief="flat",
                  font=("Segoe UI", 10), cursor="hand2",
                  padx=12, pady=7).pack(side="right", padx=(0, 8))

    def _section(self, parent, label, initial_list):
        tk.Label(parent, text=label, bg=BG, fg=FBC_DARK,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(8, 2))
        txt = tk.Text(parent, height=4, font=("Segoe UI", 9),
                      relief="flat", bg=WHITE,
                      highlightbackground=FBC_MID, highlightthickness=1,
                      wrap="none")
        txt.insert("1.0", "\n".join(initial_list))
        txt.pack(fill="x", pady=(0, 4))
        return txt

    def _parse(self, widget):
        raw = widget.get("1.0", "end").strip()
        return [ln.strip() for ln in raw.splitlines() if ln.strip()]

    def _save(self):
        to_result = self._parse(self._to_frame) if self._has_to else None
        cc_result = self._parse(self._cc_frame)
        self._on_save(to_result, cc_result)
        self.destroy()


# ════════════════════════════════════════════════════════════════════════════
#  ── SARESTOCK LOGIC ────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

EO_HEADERS = [
    "Exchange","Market","Symbol","Buy/Sell","Participant","Custodian","Client",
    "Trader","Short Sell","Price","Volume","Yield %","Accrued Interest","Order No.",
    "Ticket No.","Date/Time","Execution Date/Time","Type","Filled Volume",
    "Remaining Volume","Disc. Volume","Trigger Price","Order Initiator","Pricing Mechanism"
]
PREVIEW_COLS = ["Exchange","Market","Participant","Custodian","Client",
                "Symbol","Buy/Sell","Price","Volume","Ticket No."]

SARESTOCK_EMAIL_SUBJECT = "DEALS CONFIRMATION"

def get_sarestock_email_body(sender_name=""):
    name = sender_name.strip() or "FBC Securities"
    return f"Good day,\r\n\r\nKindly find attached for deals confirmation.\r\n\r\nRegards,\r\n{name}."

FIELD_MAP = [
    ("Security","Symbol"),("SCA Code","Custodian"),("Buy/Sell","Buy/Sell"),
    ("Quantity","Volume + Filled Vol."),("Price","Yield"),
    ("Ticket No.","Match Reference"),("Trader","Trader + Order Init."),
    ("VFX → VFEX","Exchange (+E)"),("VFEX = FBCSZWVX","Participant (fixed)"),
    ("ZSE = FBCSZWHX","Participant (fixed)"),("…-02 → …-0002","Client (zero-pad)"),
    ("DD/MM/YYYY …","Date/Time (auto)"),
]

def get_exchange(market):
    u=(market or "").strip().upper(); return "VFEX" if u=="VFX" else (u or "ZSE")

def get_participant(exch): return "FBCSZWVX" if exch=="VFEX" else "FBCSZWHX"

def get_market(sym,exch):
    s=(sym or "").upper().strip()
    if exch=="VFEX" or s.endswith(".VX"): return "REG"
    if s.endswith(".ZW"):
        if any(r in s for r in ["FHML","ZMRE","STFL","IPFL","HAFP","REVH"]): return "REIT"
        if any(o in s for o in ["SEED","CFI","CAFCA"]): return "ODD"
        return "REG"
    return "REG"

def pad_client(c):
    s=str(c or "").strip(); d=s.rfind("-")
    return s if d==-1 else f"{s[:d]}-{s[d+1:].zfill(4)}"

def get_now():
    d=datetime.now()
    return f"{d.day}/{d.month}/{d.year} {d.hour}:{d.minute:02d}"

def stamp():
    d=datetime.now(); return f"{d.day}_{d.month}_{d.year}"

def transform_rows(raw_rows):
    now = get_now()
    out = []
    for r in raw_rows:
        exch = get_exchange(r.get("Market", ""))
        sym  = r.get("Security", "")
        out.append({
            "Exchange": exch, "Market": get_market(sym, exch), "Symbol": sym,
            "Buy/Sell": r.get("Buy/Sell", ""), "Participant": get_participant(exch),
            "Custodian": r.get("SCA Code", ""), "Client": pad_client(r.get("CSD Account", "")),
            "Trader": r.get("Trader", ""), "Short Sell": "NO",
            "Price": r.get("Yield", ""), "Volume": r.get("Quantity", ""),
            "Yield %": "0", "Accrued Interest": "0", "Order No.": r.get("Trade Leg", "").lstrip("0") or "0",
            "Ticket No.": r.get("Match Reference", "").lstrip("0") or "0",
            "Date/Time": now, "Execution Date/Time": now,
            "Type": "Limit", "Filled Volume": r.get("Quantity", ""),
            "Remaining Volume": "0", "Disc. Volume": "0", "Trigger Price": "0",
            "Order Initiator": r.get("Trader", ""), "Pricing Mechanism": ""
        })
    return out, now

def generate_csv(rows, out_dir, label):
    label = label.upper()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"ExportExecutedOrders_{label}_{ts}.csv")
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=EO_HEADERS)
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in EO_HEADERS})
    return path

def generate_matched_excel(source_path, raw_rows, out_dir):
    exch = get_exchange(raw_rows[0].get("Market", "")) if raw_rows else ""
    label = "VFEX" if exch == "VFEX" else "ZSE"
    ext = os.path.splitext(source_path)[1] if source_path else ".xlsx"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(out_dir, f"MATCHED TRADES, {label}_{ts}{ext}")
    src_norm = os.path.normcase(os.path.abspath(source_path))
    dst_norm = os.path.normcase(os.path.abspath(dest))
    if src_norm == dst_norm:
        return dest
    try:
        _shutil.copy2(source_path, dest)
    except PermissionError:
        raise PermissionError(
            f"Could not copy '{os.path.basename(source_path)}' — "
            "please close it in Excel and try again.")
    return dest
ANESU_COLUMNS = ["Market", "CSD Account", "SCA Code", "Name",
                  "Security", "Buy/Sell", "Quantity", "Yield"]

def generate_anesu_excel(raw_rows, out_dir):
    """Trimmed excel — raw columns only, no transformation. Font: Aptos Narrow."""
    _require("pandas")
    import pandas as pd
    from openpyxl.styles import Font
    exch_raw = (raw_rows[0].get("Market", "") or "").strip().upper() if raw_rows else ""
    label = "VFEX" if exch_raw in ("VFX", "VFEX") else "ZSE"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"CONFIRMATION TEMPLATE, {label}_{ts}.xlsx")
    trimmed = [{col: r.get(col, "") for col in ANESU_COLUMNS} for r in raw_rows]
    df = pd.DataFrame(trimmed, columns=ANESU_COLUMNS)
    df.to_excel(path, index=False)

    wb = _require("openpyxl").load_workbook(path)
    ws = wb.active
    font = Font(name="Aptos Narrow", size=11)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    wb.save(path)
    return path
def generate_pdf(raw_rows, raw_headers, out_dir):
    _require("fpdf", "fpdf2")
    from fpdf import FPDF

    FONT_SIZE  = 6.5
    LINE_H     = 5.0
    HEADER_H   = 6.5
    CHAR_W     = FONT_SIZE * 0.50
    MAX_CHARS  = 28
    MIN_COL_MM = 8
    MAX_COL_MM = 55
    MARGIN     = 8
    PAGE_H_MM  = 210

    def _safe(text):
        return str(text).encode("latin-1", errors="replace").decode("latin-1")

    exch_raw = (raw_rows[0].get("Market", "") or "").strip().upper() if raw_rows else ""
    exch = "VFEX" if exch_raw in ("VFX", "VFEX") else "ZSE"
    out_path = os.path.join(out_dir, f"MATCHED TRADES, {exch}.pdf")

    def _col_w(hdr, rows, key):
        mx = len(str(hdr))
        for r in rows:
            mx = max(mx, len(str(r.get(key, "") or "")))
        mx = min(mx, MAX_CHARS)
        return max(MIN_COL_MM, min(MAX_COL_MM, mx * CHAR_W))

    col_widths = [_col_w(h, raw_rows, h) for h in raw_headers]
    total_content_w = sum(col_widths)
    page_w = total_content_w + 2 * MARGIN
    ts_label   = datetime.now().strftime("%d %b %Y  %H:%M")
    total_rows = len(raw_rows)

    pdf = FPDF(orientation="L", unit="mm", format=(PAGE_H_MM, page_w))
    pdf.set_margins(MARGIN, MARGIN, MARGIN)
    pdf.set_auto_page_break(auto=True, margin=14)

    def _page_header():
        pdf.set_font("Courier", style="B", size=6.5)
        pdf.set_text_color(90, 90, 90)
        label = _safe(
            f"MATCHED TRADES - {exch}  |  "
            f"{total_rows} row(s)  |  {ts_label}  |  "
            f"All {len(raw_headers)} columns — scroll right to see full table")
        pdf.cell(0, 4.5, label, border=0, align="L")
        pdf.ln(6)
        pdf.set_text_color(0, 0, 0)

    def _draw_col_headers():
        pdf.set_font("Courier", style="B", size=FONT_SIZE)
        pdf.set_text_color(0, 0, 0)
        for i, h in enumerate(raw_headers):
            pdf.cell(col_widths[i], HEADER_H, _safe(str(h)[:MAX_CHARS]), border=0, align="L")
        pdf.ln()
        y = pdf.get_y()
        pdf.set_draw_color(150, 150, 150)
        pdf.line(MARGIN, y, page_w - MARGIN, y)
        pdf.set_draw_color(0, 0, 0)
        pdf.ln(0.8)

    def _draw_data_rows():
        pdf.set_font("Courier", size=FONT_SIZE)
        pdf.set_text_color(0, 0, 0)
        for row in raw_rows:
            if pdf.get_y() > pdf.h - 16:
                pdf.add_page()
                _page_header()
                _draw_col_headers()
                pdf.set_font("Courier", size=FONT_SIZE)
                pdf.set_text_color(0, 0, 0)
            for i, h in enumerate(raw_headers):
                val = _safe(str(row.get(h, "") or "")[:MAX_CHARS])
                pdf.cell(col_widths[i], LINE_H, val, border=0, align="L")
            pdf.ln()

    pdf.add_page()
    _page_header()
    _draw_col_headers()
    _draw_data_rows()
    pdf.output(out_path)
    return out_path


def open_sarestock_outlook(file_paths, sender_name=""):
    """Open Outlook deals-confirmation email using saved To/CC recipients."""
    _require("win32com.client","pywin32")
    import win32com.client as win32
    to_list, cc_list = load_sarestock_recipients()
    outlook = win32.Dispatch("outlook.application")
    mail    = outlook.CreateItem(0)
    mail.Subject = SARESTOCK_EMAIL_SUBJECT
    mail.Body    = get_sarestock_email_body(sender_name)
    mail.To  = "; ".join(to_list)
    mail.CC  = "; ".join(cc_list)
    for fp in file_paths:
        if fp and os.path.exists(fp): mail.Attachments.Add(fp)
    mail.Display(True)


# ════════════════════════════════════════════════════════════════════════════
#  ── EMAILER LOGIC ──────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════
CONTACTS_FILE    = os.path.join(os.path.expanduser("~"),".fbc_dealnote_contacts.json")
SENDER_NAME_FILE = os.path.join(os.path.expanduser("~"),".fbc_sender_name.txt")

def load_sender_name():
    try:
        with open(SENDER_NAME_FILE) as f:
            return f.read().strip()
    except Exception:
        return ""

def save_sender_name(name):
    with open(SENDER_NAME_FILE,"w") as f:
        f.write(name.strip())

SYNC_CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".fbc_sync_config.json")
SHEET_WORKSHEET  = "Contacts"

def _load_sync_config():
    try:
        with open(SYNC_CONFIG_FILE) as f:
            return json.load(f)
    except Exception:
        return {"sheet_id": "", "service_account_path": ""}

def _save_sync_config(cfg):
    with open(SYNC_CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

def _get_gsheet():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return None, "gspread not installed. Run: pip install gspread google-auth"
    cfg = _load_sync_config()
    sa_path  = cfg.get("service_account_path", "")
    sheet_id = cfg.get("sheet_id", "")
    if not sa_path or not sheet_id:
        return None, "sync_not_configured"
    if not os.path.exists(sa_path):
        return None, f"Service account file not found:\n{sa_path}"
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds  = Credentials.from_service_account_file(sa_path, scopes=scopes)
        gc     = gspread.authorize(creds)
        sh     = gc.open_by_key(sheet_id)
        try:
            ws = sh.worksheet(SHEET_WORKSHEET)
        except Exception:
            ws = sh.add_worksheet(title=SHEET_WORKSHEET, rows=500, cols=2)
            ws.append_row(["Name", "Email"])
        return ws, None
    except Exception as e:
        return None, str(e)

def push_contacts_to_sheet(contacts):
    ws, err = _get_gsheet()
    if ws is None:
        return False, err
    try:
        rows = [["Name", "Email"]]
        for name, data in sorted(contacts.items()):
            rows.append([name, data.get("email", "")])
        ws.clear()
        ws.update(range_name="A1", values=rows)
        return True, f"Synced {len(contacts)} contacts to Google Sheets"
    except Exception as e:
        return False, str(e)

def pull_contacts_from_sheet():
    ws, err = _get_gsheet()
    if ws is None:
        return None, err
    try:
        all_rows = ws.get_all_values()
        if not all_rows:
            return {}, "Sheet is empty"
        data_rows = all_rows[1:] if all_rows[0] == ["Name", "Email"] else all_rows
        contacts = {}
        for row in data_rows:
            if len(row) >= 1 and row[0].strip():
                name  = row[0].strip().upper()
                email = row[1].strip() if len(row) > 1 else ""
                contacts[name] = {"email": email}
        return contacts, f"Loaded {len(contacts)} contacts from Google Sheets"
    except Exception as e:
        return None, str(e)

KNOWN_CUSTODIANS = ["FBCZSEZW","CBZCZWHX","STINZWVX","CBCZSEZW","FBCSZWVX"]

CUSTODIAN_PREFIX_MAP = [
    ("FBC","FBCZSEZW"),("CBC","CBCZSEZW"),("CBZ","CBZCZWHX"),
    ("STIN","STINZWVX"),("STIZ","STINZWVX"),
]

_FBC_CC = [
    "Manatsa Tagwireyi <Manatsa.Tagwireyi@fbc.co.zw>",
    "Norman Chirima <Norman.Chirima@fbc.co.zw>",
    "Enock Rukarwa <Enock.Rukarwa@fbc.co.zw>",
    "Richard Mashava <Richard.Mashava@fbc.co.zw>",
    "Anesu Zingundu <Anesu.Zingundu@fbc.co.zw>",
]

# Default routing (used when no override is saved for a custodian)
CUSTODIAN_ROUTING = {
    "FBCZSEZW":{"label":"FBC Securities (ZSE)",
        "to":["Faith Chikati <Faith.Chikati@fbc.co.zw>"],
        "cc":["Custodial Services <CustodialServices@fbc.co.zw>"]+_FBC_CC},
    "CBZCZWHX":{"label":"CBZ (ZSE)",
        "to":["Sharleen Kapininga <skapininga@cbz.co.zw>","Phillipa Gurure <pgurure@cbz.co.zw>"],
        "cc":["Custodial Services <custodialservices@cbz.co.zw>"]+_FBC_CC},
    "STINZWVX":{"label":"Stanbic",
        "to":["Maigurira, Debra D <maigurirad@stanbic.com>","Chibvongodze, Kudakwashe K <chibvongodzek@stanbic.com>"],
        "cc":["custodyzim <custodyzim@standardbank.co.za>"]+_FBC_CC},
    "CBCZSEZW":{"label":"CABS / Old Mutual",
        "to":["Darlington Tatenda Maenda <darlingtonm@oldmutual.co.zw>"],
        "cc":["Custodial Services Division <custodialservicesdivision@cabs.co.zw>"]+_FBC_CC},
    "FBCSZWVX":{"label":"FBC Securities (VFEX)",
        "to":["Faith Chikati <Faith.Chikati@fbc.co.zw>"],
        "cc":["Custodial Services <CustodialServices@fbc.co.zw>"]+_FBC_CC},
}

def get_effective_custodian_routing(code):
    """Return routing dict merging saved overrides on top of defaults."""
    base     = CUSTODIAN_ROUTING.get(code, {})
    override = load_custodian_overrides().get(code, {})
    if not base:
        return None
    return {
        "label": base["label"],
        "to":    override.get("to", base["to"]),
        "cc":    override.get("cc", base["cc"]),
    }

def get_custodian_body(multi=False, sender_name=""):
    name = sender_name.strip() or "FBC Securities"
    if multi:
        return f"Good day,\r\n\r\nKindly find attached today's deal notes.\r\n\r\nRegards,\r\n{name}."
    return f"Good day,\r\n\r\nKindly find attached today's deal note.\r\n\r\nRegards,\r\n{name}."

def get_client_body(client, multi=False, sender_name=""):
    name = sender_name.strip() or "FBC Securities"
    if multi:
        return f"Dear {client},\r\n\r\nPlease find attached your deal notes for today's transactions.\r\n\r\nRegards,\r\n{name}."
    return f"Dear {client},\r\n\r\nPlease find attached your deal note for today's transaction.\r\n\r\nRegards,\r\n{name}."

def _name_tokens(name):
    return frozenset(w.strip() for w in name.upper().split() if w.strip())

def _names_match(a, b):
    ta, tb = _name_tokens(a), _name_tokens(b)
    if not ta or not tb:
        return False
    if ta == tb or ta.issubset(tb) or tb.issubset(ta):
        return True
    return False

def find_contact(contacts, client_name):
    if client_name in contacts:
        return contacts[client_name]
    for saved, data in contacts.items():
        if _names_match(client_name, saved):
            return data
    for saved, data in contacts.items():
        if client_name.startswith(saved) or saved.startswith(client_name):
            return data
    return {}

def load_contacts():
    try:
        with open(CONTACTS_FILE) as f:
            local = json.load(f)
    except Exception:
        local = {}
    cfg = _load_sync_config()
    if cfg.get("sheet_id") and cfg.get("service_account_path"):
        try:
            sheet_contacts, err = pull_contacts_from_sheet()
            if sheet_contacts is not None and sheet_contacts:
                merged = dict(local)
                merged.update(sheet_contacts)
                save_contacts(merged)
                return merged
        except Exception:
            pass
    return local

def save_contacts(data):
    with open(CONTACTS_FILE,"w") as f: json.dump(data,f,indent=2)

def parse_client_name_from_filename(fname):
    base = os.path.splitext(fname)[0]
    base = base.replace("_", " ").strip()
    base = re.sub(r'\s*\(\d+\)\s*$', '', base)
    base = re.sub(r'\s*_\d{6,}\s*$', '', base)
    base = re.sub(r',\s*[\d,]+\s+[A-Z]{2,6}\.?\s*$', '', base)
    base = re.sub(r'\s+\d{6,}\s*$', '', base)
    # ADD THIS: strip anything after a comma (address suffixes like ", PADE 123")
    if ',' in base:
        base = base.split(',')[0].strip()
    return base.strip().upper()

def parse_client_name_from_pdf(pdf_path):
    SKIP_WORDS = [
        "FBC SECURITIES", "CONTRACT NOTE", "FISCAL TAX", "INVOICE",
        "CHARGES", "RATES TABLE", "SETTLEMENT", "EXCHANGE", "ZIMBABWE STOCK",
        "MEMBERS OF", "FOR AND ON", "VERIFICATION", "NO OF SHARES",
        "DESCRIPTION", "CONSIDERATION", "THIS CONTRACT", "DEAL DATE",
        "DEAL NUMBER", "CSD CODE", "CUSTODIAL", "SUBJECT TO",
        "ECONET", "INFRA", "FIRST BANKING", "PFUMA", "ISIN",
        "SUB TOTAL", "VAT TOTAL", "INVOICE TOTAL", "76 S.", "TEL:", "VAT:",
    ]
    def _is_valid(c):
        c = c.strip().upper()
        if len(c) < 4: return False
        if re.match(r'^[\d\s/\-\.,%]+$', c): return False
        if any(skip in c for skip in SKIP_WORDS): return False
        return True
    try:
        import fitz
        doc  = fitz.open(pdf_path)
        text = doc[0].get_text()
        doc.close()
        m = re.search(r'Deal Date\s+[\d/\-]+\s*\n\s*(.+)', text)
        if m:
            candidate = m.group(1).strip().upper()
            if _is_valid(candidate): return candidate
        m = re.search(r'Fiscal Tax Invoice\s*\n+([^\n]{4,80})\n', text, re.IGNORECASE)
        if m:
            candidate = m.group(1).strip().upper()
            if _is_valid(candidate) and not re.search(r'\d{4}/\d{2}/\d{2}', candidate):
                return candidate
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        for line in lines:
            upper = line.upper()
            if (len(upper.split()) >= 2
                    and re.match(r'^[A-Z][A-Z\s]+$', upper)
                    and len(upper) > 8
                    and _is_valid(upper)):
                return upper
    except Exception:
        pass
    return None
def parse_custodian_from_pdf(pdf_path):
    try:
        import fitz
        doc=fitz.open(pdf_path); text="".join(p.get_text() for p in doc); doc.close()
        for code in KNOWN_CUSTODIANS:
            if code in text: return code
        candidates=re.findall(r'\b([A-Z]{4,10})\b',text)
        for c in candidates:
            for prefix,canonical in CUSTODIAN_PREFIX_MAP:
                if c.startswith(prefix): return canonical
    except Exception: pass
    return None

def parse_deal_info_from_pdf(pdf_path):
    info={"deal_number":"","counter":"","deal_date":""}
    try:
        import fitz
        doc=fitz.open(pdf_path); text="".join(p.get_text() for p in doc); doc.close()
        m=re.search(r'Deal Number\s+(\d+)',text)
        if m: info["deal_number"]=m.group(1)
        m=re.search(r'Deal Date\s+([\d/]+)',text)
        if m: info["deal_date"]=m.group(1)
        m=re.search(r'\b([A-Z]{2,6}\.ZW|[A-Z]{2,6}\.VX)\b',text)
        if m: info["counter"]=m.group(1)
    except Exception: pass
    return info

def open_outlook(to_list,cc_list,subject,body,attachments):
    try:
        import win32com.client as win32
        outlook=win32.Dispatch("outlook.application"); mail=outlook.CreateItem(0)
        mail.To="; ".join(to_list); mail.CC="; ".join(cc_list)
        mail.Subject=subject; mail.Body=body
        for path in attachments:
            if os.path.exists(path): mail.Attachments.Add(path)
        mail.Display(True)
    except ImportError: raise ImportError("pywin32 not installed.\n\nRun:  pip install pywin32")
    except Exception as e: raise RuntimeError(f"Outlook error: {e}")


# ════════════════════════════════════════════════════════════════════════════
#  CONTACTS DIALOG
# ════════════════════════════════════════════════════════════════════════════
class ContactsDialog(tk.Toplevel):
    def __init__(self,parent,contacts,on_save):
        super().__init__(parent)
        self.title("Manage Client Contacts"); self.geometry("740x560")
        self.configure(bg=BG); self.contacts=dict(contacts)
        self.on_save=on_save; self._current_name=None
        self.grab_set(); self._build()

    def _build(self):
        hdr=tk.Frame(self,bg=FBC_DARK,pady=10,padx=14); hdr.pack(fill="x")
        tk.Label(hdr,text="👥  Manage Client Contacts",bg=FBC_DARK,fg=WHITE,
                 font=("Segoe UI",11,"bold")).pack(side="left")
        tk.Label(hdr,text=f"  {len(self.contacts)} clients saved",bg=FBC_DARK,fg="#90CAF9",
                 font=("Segoe UI",9)).pack(side="left",padx=8)
        tk.Button(hdr,text="☁ Setup Sync",command=self._setup_sync,
                  bg=FBC_ACCENT,fg=WHITE,relief="flat",font=("Segoe UI",8,"bold"),
                  cursor="hand2",padx=8,pady=3).pack(side="right",padx=(0,4))
        tk.Button(hdr,text="⬆ Push to Sheet",command=self._push_to_sheet,
                  bg="#1A6B3A",fg=WHITE,relief="flat",font=("Segoe UI",8,"bold"),
                  cursor="hand2",padx=8,pady=3).pack(side="right",padx=(0,4))
        tk.Button(hdr,text="⬇ Pull from Sheet",command=self._pull_from_sheet,
                  bg="#1A3A6B",fg=WHITE,relief="flat",font=("Segoe UI",8,"bold"),
                  cursor="hand2",padx=8,pady=3).pack(side="right",padx=(0,8))

        body=tk.Frame(self,bg=BG); body.pack(fill="both",expand=True,padx=12,pady=10)
        left=tk.Frame(body,bg=WHITE,relief="flat",bd=1); left.pack(side="left",fill="y",padx=(0,8))
        tk.Label(left,text="Clients",bg=FBC_MID,fg=WHITE,
                 font=("Segoe UI",9,"bold"),pady=6,padx=8).pack(fill="x")
        s_frame=tk.Frame(left,bg=WHITE,padx=4,pady=4); s_frame.pack(fill="x")
        tk.Label(s_frame,text="🔍",bg=WHITE,font=("Segoe UI",10)).pack(side="left")
        self.search_var=tk.StringVar()
        self.search_var.trace_add("write",lambda *_:self._filter_list())
        search_entry=tk.Entry(s_frame,textvariable=self.search_var,
                              font=("Segoe UI",9),relief="flat",bd=0,
                              bg="#F0F4F8",width=20)
        search_entry.pack(side="left",fill="x",expand=True,padx=4)
        tk.Button(s_frame,text="✕",command=lambda:(self.search_var.set(""),search_entry.focus()),
                  bg=WHITE,fg="#8096B0",relief="flat",font=("Segoe UI",8),
                  cursor="hand2",padx=2).pack(side="left")
        lb_frame=tk.Frame(left,bg=WHITE); lb_frame.pack(fill="both",expand=True,padx=4,pady=(0,4))
        self.listbox=tk.Listbox(lb_frame,width=26,font=("Segoe UI",9),
                                selectbackground=FBC_MID,activestyle="none",
                                relief="flat",bd=0)
        lb_sb=ttk.Scrollbar(lb_frame,orient="vertical",command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=lb_sb.set)
        self.listbox.pack(side="left",fill="both",expand=True)
        lb_sb.pack(side="right",fill="y")
        self.listbox.bind("<<ListboxSelect>>",self._on_select)
        br=tk.Frame(left,bg=WHITE); br.pack(fill="x",padx=4,pady=(0,6))
        tk.Button(br,text="+ Add",command=self._add,bg=GREEN_DARK,fg=WHITE,
                  relief="flat",font=("Segoe UI",8,"bold"),cursor="hand2").pack(side="left",padx=(0,4))
        tk.Button(br,text="✕ Delete",command=self._delete,bg=RED_DARK,fg=WHITE,
                  relief="flat",font=("Segoe UI",8,"bold"),cursor="hand2").pack(side="left")
        right=tk.Frame(body,bg=WHITE,relief="flat",bd=1); right.pack(side="left",fill="both",expand=True)
        tk.Label(right,text="Contact Details",bg=FBC_MID,fg=WHITE,
                 font=("Segoe UI",9,"bold"),pady=6,padx=8).pack(fill="x")
        self.detail=tk.Frame(right,bg=WHITE); self.detail.pack(fill="both",expand=True,padx=14,pady=12)
        self._show_detail(None)
        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",padx=12,pady=(0,10))
        tk.Button(bot,text="💾  Save & Close",command=self._save,bg=FBC_MID,fg=WHITE,
                  font=("Segoe UI",10,"bold"),relief="flat",padx=16,pady=8,cursor="hand2").pack(side="right")
        self._filter_list()

    def _filter_list(self):
        term=self.search_var.get().strip().upper()
        self.listbox.delete(0,tk.END)
        for n in sorted(self.contacts):
            if term in n.upper():
                self.listbox.insert(tk.END,n)
        if self._current_name:
            for i in range(self.listbox.size()):
                if self.listbox.get(i)==self._current_name:
                    self.listbox.selection_set(i); break

    def _refresh_list(self): self._filter_list()

    def _on_select(self,_=None):
        sel=self.listbox.curselection()
        if sel:
            self._current_name=self.listbox.get(sel[0])
            self._show_detail(self._current_name)

    def _show_detail(self,name):
        for w in self.detail.winfo_children(): w.destroy()
        if not name:
            tk.Label(self.detail,text="Select a client on the left\nto view or edit their details.",
                     bg=WHITE,fg="#8096B0",font=("Segoe UI",9),justify="center").pack(pady=30); return
        data=self.contacts.get(name,{"email":""})
        tk.Label(self.detail,text="Client Name:",bg=WHITE,fg="#607080",
                 font=("Segoe UI",8,"bold")).pack(anchor="w")
        name_row=tk.Frame(self.detail,bg=WHITE); name_row.pack(fill="x",pady=(2,10))
        self.entry_name=tk.Entry(name_row,font=("Segoe UI",10),width=34)
        self.entry_name.insert(0,name)
        self.entry_name.pack(side="left")
        tk.Button(name_row,text="✏ Rename",command=lambda n=name:self._rename(n),
                  bg=FBC_MID,fg=WHITE,relief="flat",font=("Segoe UI",8,"bold"),
                  cursor="hand2",padx=8,pady=4).pack(side="left",padx=6)
        tk.Label(self.detail,text="Client Email:",bg=WHITE,fg="#607080",
                 font=("Segoe UI",8,"bold")).pack(anchor="w")
        self.entry_email=tk.Entry(self.detail,font=("Segoe UI",10),width=42)
        self.entry_email.insert(0,data.get("email",""))
        self.entry_email.pack(anchor="w",pady=(2,4))
        saved=data.get("email","")
        hint_txt="No email saved yet" if not saved else f"Saved: {saved}"
        hint_col=RED_DARK if not saved else GREEN_DARK
        tk.Label(self.detail,text=hint_txt,bg=WHITE,fg=hint_col,font=("Segoe UI",8)).pack(anchor="w",pady=(0,12))
        tk.Button(self.detail,text="✔  Apply Email",command=lambda n=name:self._apply(n),
                  bg=GREEN_DARK,fg=WHITE,relief="flat",font=("Segoe UI",9,"bold"),
                  cursor="hand2",padx=12,pady=6).pack(anchor="w")

    def _rename(self,old_name):
        new_name=self.entry_name.get().strip().upper()
        if not new_name:
            messagebox.showwarning("Empty","Name cannot be empty.",parent=self); return
        if new_name==old_name:
            messagebox.showinfo("No Change","Name is the same.",parent=self); return
        if new_name in self.contacts:
            messagebox.showwarning("Duplicate",f"'{new_name}' already exists.",parent=self); return
        self.contacts[new_name]=self.contacts.pop(old_name)
        self._current_name=new_name
        self._filter_list(); self._show_detail(new_name)
        messagebox.showinfo("Renamed",f"'{old_name}' renamed.\n\nClick 'Save & Close' to keep this change.",parent=self)

    def _apply(self,name):
        self.contacts[name]={"email":self.entry_email.get().strip()}
        self._show_detail(name)
        messagebox.showinfo("Saved",f"Email saved for {name}.\nClick 'Save & Close' to write to disk.",parent=self)

    def _add(self):
        dlg=tk.Toplevel(self); dlg.title("Add Client")
        dlg.geometry("340x130"); dlg.configure(bg=BG); dlg.grab_set()
        tk.Label(dlg,text="Client name (as it appears in the filename):",
                 bg=BG,font=("Segoe UI",9)).pack(pady=(14,4),padx=12)
        e=tk.Entry(dlg,font=("Segoe UI",10),width=34); e.pack(padx=12); e.focus()
        def ok():
            n=e.get().strip().upper()
            if not n: return
            if n in self.contacts:
                messagebox.showwarning("Duplicate",f"'{n}' already exists.",parent=dlg); return
            self.contacts[n]={"email":""}
            self._current_name=n
            self._filter_list(); dlg.destroy(); self._show_detail(n)
        e.bind("<Return>",lambda _:ok())
        tk.Button(dlg,text="Add",command=ok,bg=FBC_MID,fg=WHITE,
                  relief="flat",font=("Segoe UI",9,"bold"),cursor="hand2",
                  padx=12,pady=6).pack(pady=10)

    def _delete(self):
        sel=self.listbox.curselection()
        if not sel: return
        name=self.listbox.get(sel[0])
        if messagebox.askyesno("Delete",f"Delete '{name}' from contacts?",parent=self):
            self.contacts.pop(name,None); self._current_name=None
            self._filter_list(); self._show_detail(None)

    def _setup_sync(self):
        cfg = _load_sync_config()
        dlg = tk.Toplevel(self); dlg.title("Setup Google Sheets Sync")
        dlg.geometry("540x280"); dlg.configure(bg=BG); dlg.grab_set()
        tk.Label(dlg,text="Google Sheets Sync Setup",bg=FBC_DARK,fg=WHITE,
                 font=("Segoe UI",11,"bold"),pady=10,padx=14).pack(fill="x")
        body=tk.Frame(dlg,bg=BG,padx=16,pady=12); body.pack(fill="both",expand=True)
        tk.Label(body,text="Google Sheet ID (from the URL):",bg=BG,
                 font=("Segoe UI",9,"bold")).pack(anchor="w")
        e_sheet=tk.Entry(body,font=("Segoe UI",9),width=60)
        e_sheet.insert(0,cfg.get("sheet_id",""))
        e_sheet.pack(anchor="w",pady=(2,10),fill="x")
        tk.Label(body,text="Service Account JSON file path:",bg=BG,
                 font=("Segoe UI",9,"bold")).pack(anchor="w")
        path_row=tk.Frame(body,bg=BG); path_row.pack(fill="x",pady=(2,4))
        e_path=tk.Entry(path_row,font=("Segoe UI",9),width=48)
        e_path.insert(0,cfg.get("service_account_path",""))
        e_path.pack(side="left",fill="x",expand=True)
        def browse():
            p=filedialog.askopenfilename(title="Select Service Account JSON",
                filetypes=[("JSON","*.json"),("All","*.*")])
            if p: e_path.delete(0,tk.END); e_path.insert(0,p)
        tk.Button(path_row,text="Browse...",command=browse,bg=FBC_MID,fg=WHITE,
                  relief="flat",font=("Segoe UI",8),cursor="hand2",padx=6).pack(side="left",padx=4)
        tk.Label(body,text="See SYNC_SETUP_GUIDE.txt for step-by-step instructions.",
                 bg=BG,fg="#607080",font=("Segoe UI",8)).pack(anchor="w",pady=(4,0))
        def save_cfg():
            _save_sync_config({"sheet_id":e_sheet.get().strip(),
                               "service_account_path":e_path.get().strip()})
            messagebox.showinfo("Saved","Sync configuration saved.",parent=dlg)
            dlg.destroy()
        tk.Button(body,text="Save Configuration",command=save_cfg,
                  bg=FBC_MID,fg=WHITE,relief="flat",font=("Segoe UI",9,"bold"),
                  cursor="hand2",padx=12,pady=6).pack(anchor="w",pady=(12,0))

    def _push_to_sheet(self):
        ok, msg = push_contacts_to_sheet(self.contacts)
        if ok: messagebox.showinfo("Sync", msg, parent=self)
        elif msg == "sync_not_configured":
            messagebox.showwarning("Not Configured",
                "Sync is not set up yet.\nClick 'Setup Sync' to configure.", parent=self)
        else: messagebox.showerror("Sync Error", msg, parent=self)

    def _pull_from_sheet(self):
        contacts, msg = pull_contacts_from_sheet()
        if contacts is None:
            if msg == "sync_not_configured":
                messagebox.showwarning("Not Configured",
                    "Sync is not set up yet.\nClick 'Setup Sync' to configure.", parent=self)
            else: messagebox.showerror("Sync Error", msg, parent=self)
            return
        self.contacts.update(contacts)
        save_contacts(self.contacts)
        self._filter_list()
        messagebox.showinfo("Pulled", msg, parent=self)

    def _save(self):
        save_contacts(self.contacts)
        self.on_save(self.contacts)
        def _push():
            ok, msg = push_contacts_to_sheet(self.contacts)
            if ok: print(f"[Sync] {msg}")
            elif msg and msg != "sync_not_configured": print(f"[Sync error] {msg}")
        threading.Thread(target=_push, daemon=True).start()
        self.destroy()


# ════════════════════════════════════════════════════════════════════════════
#  SARESTOCK PAGE
# ════════════════════════════════════════════════════════════════════════════
class SarestockPage(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=BG)
        self.raw_rows = [];
        self.raw_headers = [];
        self.conv_rows = []
        self.source_path = None;
        self.gen_csv = self.gen_pdf = self.gen_mt_xlsx = self.gen_anesu_xlsx = None
        self.raw_rows2 = [];
        self.raw_headers2 = [];
        self.conv_rows2 = []
        self.source_path2 = None;
        self.gen_csv2 = self.gen_pdf2 = self.gen_mt_xlsx2 = self.gen_anesu_xlsx2 = None
        self.out_dir=os.path.join(os.path.expanduser("~"),"Downloads")
        self._build()

    def _build(self):
        info=tk.Frame(self,bg=FBC_MID,padx=16,pady=8); info.pack(fill="x")
        tk.Label(info,text="📊  Sarestock Upload Converter",bg=FBC_MID,fg=WHITE,
                 font=("Segoe UI",11,"bold")).pack(side="left")
        self.paned=tk.PanedWindow(self,orient="horizontal",bg=SEP_CLR,sashwidth=4,sashrelief="flat")
        self.paned.pack(fill="both",expand=True)
        self.left_frame,self.left_canvas,self.left_body=self._scroll_pane(self.paned)
        self.paned.add(self.left_frame,stretch="always")
        self.right_frame,self.right_canvas,self.right_body=self._scroll_pane(self.paned)
        self.paned.add(self.right_frame,stretch="always")

        self._build_bottom_bar()
        self._build_left_column()
        self._build_right_column()

    def _scroll_pane(self,parent):
        frame=tk.Frame(parent,bg=BG)
        canvas=tk.Canvas(frame,bg=BG,highlightthickness=0)
        vsb=ttk.Scrollbar(frame,orient="vertical",command=canvas.yview)
        inner=tk.Frame(canvas,bg=BG)
        inner.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0),window=inner,anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left",fill="both",expand=True); vsb.pack(side="right",fill="y")
        canvas.bind("<Enter>",lambda e,c=canvas:c.bind_all("<MouseWheel>",
            lambda ev:c.yview_scroll(-1*(ev.delta//120),"units")))
        canvas.bind("<Leave>",lambda e,c=canvas:c.unbind_all("<MouseWheel>"))
        return frame,canvas,inner

    def _build_bottom_bar(self):
        bar=tk.Frame(self,bg=BOTTOM,pady=10); bar.pack(fill="x",side="bottom")
        path_row=tk.Frame(bar,bg=BOTTOM); path_row.pack(fill="x",padx=16,pady=(0,6))
        tk.Label(path_row,text="Files saved to:",bg=BOTTOM,fg="#8BAAC8",
                 font=("Segoe UI",8)).pack(side="left")
        self.lbl_outdir=tk.Label(path_row,text=self.out_dir,bg=BOTTOM,fg="#90CAF9",
                                  font=("Segoe UI",8)); self.lbl_outdir.pack(side="left",padx=6)
        tk.Button(path_row,text="Change...",command=self._pick_outdir,bg="#1A3A6B",fg="#90CAF9",
                  relief="flat",font=("Segoe UI",8),cursor="hand2",padx=6,pady=2).pack(side="left")

        # ── Configure Deals Confirmation Recipients button ──────────────────
        tk.Button(path_row, text="⚙ Configure Recipients",
                  command=self._configure_recipients,
                  bg=FBC_ACCENT, fg=WHITE, relief="flat",
                  font=("Segoe UI", 8, "bold"), cursor="hand2",
                  padx=8, pady=2).pack(side="left", padx=(12, 0))

        tk.Button(path_row, text="Clear Uploads", command=self._clear_uploads,
                  bg=RED_DARK, fg=WHITE, relief="flat",
                  font=("Segoe UI", 8, "bold"), cursor="hand2",
                  padx=8, pady=2).pack(side="right")

        btn_row=tk.Frame(bar,bg=BOTTOM); btn_row.pack(fill="x",padx=16)
        btn_row.columnconfigure(0,weight=1); btn_row.columnconfigure(1,weight=1); btn_row.columnconfigure(2,weight=2)
        self.btn_email=tk.Button(btn_row,text="Send — ZSE Only",command=self._send_email,
            bg=GREEN_DARK,fg=WHITE,font=("Segoe UI",10,"bold"),relief="flat",pady=9,
            cursor="hand2",state="disabled")
        self.btn_email.grid(row=0,column=0,sticky="ew",padx=(0,6))
        self.btn_email2=tk.Button(btn_row,text="Send — VFEX Only",command=self._send_email2,
            bg="#1A3A6B",fg=WHITE,font=("Segoe UI",10,"bold"),relief="flat",pady=9,
            cursor="hand2",state="disabled")
        self.btn_email2.grid(row=0,column=1,sticky="ew",padx=(0,6))
        self.btn_email_both=tk.Button(btn_row,text="Send BOTH ZSE + VFEX in One Email",
            command=self._send_email_both,bg=FBC_MID,fg=WHITE,font=("Segoe UI",11,"bold"),
            relief="flat",pady=9,cursor="hand2",state="disabled")
        self.btn_email_both.grid(row=0,column=2,sticky="ew")

        # Recipient summary label (live feedback of current To/CC)
        self.lbl_recip_summary = tk.Label(bar, text="", bg=BOTTOM, fg="#5D7A99",
                                          font=("Segoe UI", 8))
        self.lbl_recip_summary.pack(pady=(4, 0))
        self._refresh_recip_summary()

    def _refresh_recip_summary(self):
        to_list, cc_list = load_sarestock_recipients()
        to_str = "; ".join(to_list) if to_list else "(none)"
        cc_str = f"{len(cc_list)} CC address{'es' if len(cc_list)!=1 else ''}"
        self.lbl_recip_summary.config(
            text=f"Deals Confirmation  →  To: {to_str}  |  {cc_str}  "
                 f"  (click ⚙ Configure Recipients to change)")

    def _configure_recipients(self):
        to_list, cc_list = load_sarestock_recipients()
        def on_save(new_to, new_cc):
            save_sarestock_recipients(new_to or [], new_cc)
            self._refresh_recip_summary()
            messagebox.showinfo("Saved",
                "Deals Confirmation recipients updated.\n\n"
                "All future sends will use the new addresses.", parent=self)
        RecipientsDialog(
            self,
            title="⚙  Deals Confirmation — Configure Recipients",
            to_list=to_list,
            cc_list=cc_list,
            on_save=on_save,
            to_label="To  (primary recipients — required)",
            cc_label="CC  (copied recipients)",
        )

    def _clear_uploads(self):
        has_data = bool(self.source_path or self.source_path2)
        if not has_data:
            messagebox.showinfo("Nothing to Clear", "No files are currently loaded.")
            return
        if not messagebox.askyesno("Clear Uploads",
                "Clear all uploaded matched trades files and start fresh?\n\n"
                "This does NOT delete any files from disk."):
            return
        self.raw_rows=[]; self.raw_headers=[]; self.conv_rows=[]
        self.source_path=None
        self.gen_csv = self.gen_pdf = self.gen_mt_xlsx = self.gen_anesu_xlsx = None
        self.raw_rows2 = [];
        self.raw_headers2 = [];
        self.conv_rows2 = []
        self.source_path2 = None
        self.gen_csv2 = self.gen_pdf2 = self.gen_mt_xlsx2 = self.gen_anesu_xlsx2 = None
        self.prev_outer1.pack_forget()
        self.prev_outer2.pack_forget()
        for w in self.info_bar1.winfo_children(): w.pack_forget()
        self.info_bar1.pack_forget()
        for w in self.info_bar2.winfo_children(): w.pack_forget()
        self.info_bar2.pack_forget()
        self.lbl_csv_done.config(text=""); self.lbl_pdf_done.config(text="")
        self.lbl_csv2_done.config(text=""); self.lbl_pdf2_done.config(text="")
        self.btn_csv.config(text="Download CSV", bg=FBC_MID, state="disabled")
        self.btn_pdf.config(text="Download PDF", bg=RED_DARK, state="disabled")
        self.btn_csv2.config(text="Download CSV", bg=FBC_MID, state="disabled")
        self.btn_pdf2.config(text="Download PDF", bg=RED_DARK, state="disabled")
        for b in (self.btn_email, self.btn_email2, self.btn_email_both):
            b.config(state="disabled")
        self.btn_email.config(text="Send — ZSE Only")
        self.btn_email2.config(text="Send — VFEX Only")
        messagebox.showinfo("Cleared", "Both upload slots cleared. Ready for a new upload.")

    def _build_left_column(self):
        p=self.left_body
        hdr=tk.Frame(p,bg=COL1_HDR); hdr.pack(fill="x",padx=12,pady=(12,0))
        tk.Label(hdr,text=" 1 ",bg=WHITE,fg=COL1_HDR,font=("Segoe UI",9,"bold"),
                 padx=4,pady=4).pack(side="left",padx=(8,0),pady=6)
        tk.Label(hdr,text="  FIRST EXCHANGE  (ZSE or VFEX)",bg=COL1_HDR,fg=WHITE,
                 font=("Segoe UI",10,"bold")).pack(side="left",pady=6)
        ucard=self._card(p,COL1_HDR)
        dz=tk.Frame(ucard,bg="#F4F8FE",relief="groove",bd=2); dz.pack(fill="x",pady=(0,10))
        inner=tk.Frame(dz,bg="#F4F8FE"); inner.pack(pady=16)
        tk.Label(inner,text="Upload Matched Trades File",bg="#F4F8FE",fg=FBC_MID,
                 font=("Segoe UI",10,"bold")).pack(pady=(4,0))
        tk.Label(inner,text=".csv or .xlsx",bg="#F4F8FE",fg="#8096B0",font=("Segoe UI",8)).pack()
        tk.Button(inner,text="  Browse...  ",command=self._pick_file,bg=FBC_MID,fg=WHITE,
                  font=("Segoe UI",10,"bold"),relief="flat",padx=14,pady=6,cursor="hand2").pack(pady=(8,0))
        self.info_bar1=tk.Frame(ucard,bg=TAG_BLUE,highlightbackground=FBC_ACCENT,highlightthickness=1)
        self.lbl_file1=tk.Label(self.info_bar1,text="",bg=TAG_BLUE,fg=FBC_DARK,font=("Segoe UI",9,"bold"))
        self.lbl_rows1=tk.Label(self.info_bar1,text="",bg=TAG_BLUE,fg=FBC_MID,font=("Consolas",8))
        self.btn_reupload1=tk.Button(self.info_bar1,text="Change",command=self._pick_file,
                                     bg=TAG_BLUE,fg=FBC_MID,relief="flat",font=("Segoe UI",8),cursor="hand2")
        dcard=self._card(p,COL1_HDR)
        tk.Label(dcard,text="DOWNLOAD",bg=CARD_BG,fg="#8096B0",font=("Segoe UI",8,"bold")).pack(anchor="w")
        btn_row=tk.Frame(dcard,bg=CARD_BG); btn_row.pack(fill="x",pady=(6,2))
        self.btn_csv=self._col_btn(btn_row,"Download CSV",self._dl_csv,FBC_MID)
        self.btn_pdf=self._col_btn(btn_row,"Download PDF",self._dl_pdf,RED_DARK)
        for b in (self.btn_csv,self.btn_pdf): b.config(state="disabled")
        self.lbl_csv_done=tk.Label(dcard,text="",bg=CARD_BG,fg=GREEN_DARK,font=("Segoe UI",8)); self.lbl_csv_done.pack(anchor="w")
        self.lbl_pdf_done=tk.Label(dcard,text="",bg=CARD_BG,fg=GREEN_DARK,font=("Segoe UI",8)); self.lbl_pdf_done.pack(anchor="w")
        self.prev_outer1=tk.Frame(p,bg=BG)
        self._build_preview_shell(self.prev_outer1,COL1_HDR,"PREVIEW — FIRST EXCHANGE","prev_body1","lbl_showing1")
        self.prev_outer1.pack_forget()

    def _build_right_column(self):
        p=self.right_body
        hdr=tk.Frame(p,bg=COL2_HDR); hdr.pack(fill="x",padx=12,pady=(12,0))
        tk.Label(hdr,text=" 2 ",bg=WHITE,fg=COL2_HDR,font=("Segoe UI",9,"bold"),
                 padx=4,pady=4).pack(side="left",padx=(8,0),pady=6)
        tk.Label(hdr,text="  SECOND EXCHANGE  (ZSE or VFEX)",bg=COL2_HDR,fg=WHITE,
                 font=("Segoe UI",10,"bold")).pack(side="left",pady=6)
        ucard=self._card(p,COL2_HDR)
        dz=tk.Frame(ucard,bg="#F4F8FE",relief="groove",bd=2); dz.pack(fill="x",pady=(0,10))
        inner=tk.Frame(dz,bg="#F4F8FE"); inner.pack(pady=16)
        tk.Label(inner,text="Upload Matched Trades File",bg="#F4F8FE",fg=FBC_MID,
                 font=("Segoe UI",10,"bold")).pack(pady=(4,0))
        tk.Label(inner,text=".csv or .xlsx",bg="#F4F8FE",fg="#8096B0",font=("Segoe UI",8)).pack()
        tk.Button(inner,text="  Browse...  ",command=self._pick_file2,bg=COL2_HDR,fg=WHITE,
                  font=("Segoe UI",10,"bold"),relief="flat",padx=14,pady=6,cursor="hand2").pack(pady=(8,0))
        self.info_bar2=tk.Frame(ucard,bg=TAG_BLUE,highlightbackground=FBC_ACCENT,highlightthickness=1)
        self.lbl_file2=tk.Label(self.info_bar2,text="",bg=TAG_BLUE,fg=FBC_DARK,font=("Segoe UI",9,"bold"))
        self.lbl_rows2=tk.Label(self.info_bar2,text="",bg=TAG_BLUE,fg=FBC_MID,font=("Consolas",8))
        self.btn_reupload2=tk.Button(self.info_bar2,text="Change",command=self._pick_file2,
                                     bg=TAG_BLUE,fg=FBC_MID,relief="flat",font=("Segoe UI",8),cursor="hand2")
        dcard=self._card(p,COL2_HDR)
        tk.Label(dcard,text="DOWNLOAD",bg=CARD_BG,fg="#8096B0",font=("Segoe UI",8,"bold")).pack(anchor="w")
        btn_row=tk.Frame(dcard,bg=CARD_BG); btn_row.pack(fill="x",pady=(6,2))
        self.btn_csv2=self._col_btn(btn_row,"Download CSV",self._dl_csv2,FBC_MID)
        self.btn_pdf2=self._col_btn(btn_row,"Download PDF",self._dl_pdf2,RED_DARK)
        for b in (self.btn_csv2,self.btn_pdf2): b.config(state="disabled")
        self.lbl_csv2_done=tk.Label(dcard,text="",bg=CARD_BG,fg=GREEN_DARK,font=("Segoe UI",8)); self.lbl_csv2_done.pack(anchor="w")
        self.lbl_pdf2_done=tk.Label(dcard,text="",bg=CARD_BG,fg=GREEN_DARK,font=("Segoe UI",8)); self.lbl_pdf2_done.pack(anchor="w")
        self.prev_outer2=tk.Frame(p,bg=BG)
        self._build_preview_shell(self.prev_outer2,COL2_HDR,"PREVIEW — SECOND EXCHANGE","prev_body2","lbl_showing2")
        self.prev_outer2.pack_forget()

    def _card(self,parent,accent):
        wrapper=tk.Frame(parent,bg=BG); wrapper.pack(fill="x",padx=12,pady=(4,0))
        tk.Frame(wrapper,bg=accent,height=2).pack(fill="x")
        body=tk.Frame(wrapper,bg=CARD_BG,padx=14,pady=12,highlightbackground=SEP_CLR,highlightthickness=1)
        body.pack(fill="x"); return body

    def _col_btn(self,parent,text,cmd,bg):
        b=tk.Button(parent,text=text,command=cmd,bg=bg,fg=WHITE,
                    font=("Segoe UI",9,"bold"),relief="flat",padx=10,pady=7,cursor="hand2")
        b.pack(side="left",padx=(0,8),pady=2); return b

    def _build_preview_shell(self,outer,color,title,body_attr,label_attr):
        outer.pack(fill="x",padx=12,pady=(4,0))
        hdr=tk.Frame(outer,bg=color); hdr.pack(fill="x")
        tk.Label(hdr,text=f"  {title}",bg=color,fg=WHITE,
                 font=("Segoe UI",9,"bold")).pack(side="left",pady=5,padx=6)
        lbl=tk.Label(hdr,text="",bg=color,fg="#90CAF9",font=("Segoe UI",8))
        lbl.pack(side="right",padx=8,pady=5); setattr(self,label_attr,lbl)
        body=tk.Frame(outer,bg=CARD_BG,padx=14,pady=10,highlightbackground=SEP_CLR,highlightthickness=1)
        body.pack(fill="x"); setattr(self,body_attr,body)


    def _build_preview(self,body_attr,label_attr,rows,tickets,now):
        body=getattr(self,body_attr); lbl=getattr(self,label_attr)
        for w in body.winfo_children(): w.destroy()
        summ=tk.Frame(body,bg=CARD_BG); summ.pack(fill="x",pady=(0,8))
        def ibox(parent,ltext,val,col):
            f=tk.Frame(parent,bg="#F0F7FF",padx=10,pady=6,highlightbackground=SEP_CLR,highlightthickness=1)
            f.grid(row=0,column=col,sticky="nsew",padx=(0,6)); parent.columnconfigure(col,weight=1)
            tk.Label(f,text=ltext,bg="#F0F7FF",fg="#8096B0",font=("Segoe UI",7,"bold")).pack(anchor="w")
            tk.Label(f,text=val,bg="#F0F7FF",fg=FBC_DARK,font=("Segoe UI",11,"bold")).pack(anchor="w")
        ibox(summ,"ROWS",str(len(rows)),0); ibox(summ,"DATE/TIME",now,1)
        ibox(summ,"TICKET RANGE",f"{tickets[0]} to {tickets[-1]}",2)
        style=ttk.Style()
        try: style.theme_use("default")
        except Exception: pass
        style.configure("Treeview.Heading",background=FBC_DARK,foreground=WHITE,relief="flat",font=("Segoe UI",8,"bold"))
        style.map("Treeview.Heading",background=[("active",FBC_MID)],foreground=[("active",WHITE)])
        style.configure("Treeview",font=("Segoe UI",8),rowheight=22)
        frm=tk.Frame(body,bg=CARD_BG); frm.pack(fill="x")
        xsb=ttk.Scrollbar(frm,orient="horizontal"); ysb=ttk.Scrollbar(frm,orient="vertical")
        tv=ttk.Treeview(frm,columns=PREVIEW_COLS,show="headings",height=min(len(rows),7),
                        xscrollcommand=xsb.set,yscrollcommand=ysb.set)
        xsb.config(command=tv.xview); ysb.config(command=tv.yview)
        for col in PREVIEW_COLS: tv.heading(col,text=col); tv.column(col,width=95,minwidth=70,anchor="w")
        for i,row in enumerate(rows):
            vals=[row.get(c,"") for c in PREVIEW_COLS]
            bs=row.get("Buy/Sell","").strip().lower()
            tag="buy" if bs=="buy" else("sell" if bs=="sell" else("even" if i%2==0 else "odd"))
            tv.insert("","end",values=vals,tags=(tag,))
        tv.tag_configure("even",background="#F7FAFF"); tv.tag_configure("odd",background=CARD_BG)
        tv.tag_configure("buy",foreground=GREEN_DARK,background="#F2FBF5")
        tv.tag_configure("sell",foreground=RED_DARK,background="#FFF5F5")
        tv.grid(row=0,column=0,sticky="nsew"); ysb.grid(row=0,column=1,sticky="ns")
        xsb.grid(row=1,column=0,sticky="ew"); frm.columnconfigure(0,weight=1)
        lbl.config(text=f"showing {len(rows)} rows")

    def _pick_file(self):
        path=filedialog.askopenfilename(title="Select First Exchange Matched Trades File",
            filetypes=[("CSV / Excel","*.csv *.xlsx *.xls"),("All files","*.*")])
        if path: self._load_file(path)

    def _load_file(self,path):
        try:
            pd=_require("pandas")
            df=pd.read_csv(path,dtype=str).fillna("") if path.lower().endswith(".csv") else pd.read_excel(path,dtype=str).fillna("")
            self.raw_headers=list(df.columns); self.raw_rows=df.to_dict("records")
            if not self.raw_rows: raise ValueError("File is empty.")
            self.source_path = path;
            self.conv_rows, now = transform_rows(self.raw_rows)
            tickets = [r.get("Ticket No.", "").lstrip("0") or "0" for r in self.conv_rows]
            self.gen_csv = self.gen_pdf = self.gen_mt_xlsx = self.gen_anesu_xlsx = None
            self.lbl_csv_done.config(text=""); self.lbl_pdf_done.config(text="")
            self.btn_csv.config(text="Download CSV",bg=FBC_MID)
            self.btn_pdf.config(text="Download PDF",bg=RED_DARK)
            fname=os.path.basename(path); exch=get_exchange(self.raw_rows[0].get("Market",""))
            for w in self.info_bar1.winfo_children(): w.pack_forget()
            self.info_bar1.pack(fill="x",pady=(0,6))
            tk.Label(self.info_bar1,text="OK",bg=TAG_BLUE,font=("Segoe UI",10)).pack(side="left",padx=(6,2),pady=4)
            self.lbl_file1.config(text=fname); self.lbl_file1.pack(side="left",pady=4)
            self.lbl_rows1.config(text=f"  {len(self.conv_rows)} rows  |  {tickets[0]} to {tickets[-1]}")
            self.lbl_rows1.pack(side="left",pady=4); self.btn_reupload1.pack(side="right",padx=6,pady=4)
            for b in (self.btn_csv,self.btn_pdf): b.config(state="normal")
            self.btn_email.config(text=f"Send — {exch} Only",state="normal")
            self.prev_outer1.pack(fill="x",padx=12,pady=(4,0))
            self._build_preview("prev_body1","lbl_showing1",self.conv_rows,tickets,now)
            if self.source_path2: self.btn_email_both.config(state="normal")
        except Exception as e: messagebox.showerror("Error loading file",str(e))

    def _dl_csv(self):
        try:
            exch=get_exchange(self.raw_rows[0].get("Market",""))
            self.gen_csv=generate_csv(self.conv_rows,self.out_dir,exch)
            self.lbl_csv_done.config(text=f"OK  {os.path.basename(self.gen_csv)} saved")
            self.btn_csv.config(text="CSV Downloaded",bg="#1B5E20")
        except Exception as e: messagebox.showerror("CSV Error",str(e))

    def _dl_pdf(self):
        try:
            self.gen_pdf=generate_pdf(self.raw_rows,self.raw_headers,self.out_dir)
            self.lbl_pdf_done.config(text=f"OK  {os.path.basename(self.gen_pdf)} saved")
            self.btn_pdf.config(text="PDF Downloaded",bg="#7B1010")
        except Exception as e: messagebox.showerror("PDF Error",str(e))

    def _ensure_email_files(self):
        if not self.gen_pdf:
            self.gen_pdf = generate_pdf(self.raw_rows, self.raw_headers, self.out_dir)
            self.lbl_pdf_done.config(text=f"OK  {os.path.basename(self.gen_pdf)} saved")
        if not self.gen_mt_xlsx:
            self.gen_mt_xlsx = generate_matched_excel(self.source_path, self.raw_rows, self.out_dir)
        if not self.gen_anesu_xlsx:
            self.gen_anesu_xlsx = generate_anesu_excel(self.raw_rows, self.out_dir)

    def _send_email(self):
        try:
            self._ensure_email_files()
            open_sarestock_outlook([self.gen_pdf, self.gen_mt_xlsx, self.gen_anesu_xlsx],
                                   sender_name=load_sender_name())
        except ImportError: messagebox.showerror("pywin32 not installed","Run:  pip install pywin32")
        except Exception as e: messagebox.showerror("Outlook Error",str(e))

    def _pick_file2(self):
        path=filedialog.askopenfilename(title="Select Second Exchange Matched Trades File",
            filetypes=[("CSV / Excel","*.csv *.xlsx *.xls"),("All files","*.*")])
        if path: self._load_file2(path)

    def _load_file2(self,path):
        try:
            pd=_require("pandas")
            df=pd.read_csv(path,dtype=str).fillna("") if path.lower().endswith(".csv") else pd.read_excel(path,dtype=str).fillna("")
            self.raw_headers2=list(df.columns); self.raw_rows2=df.to_dict("records")
            if not self.raw_rows2: raise ValueError("File is empty.")
            self.source_path2 = path;
            self.conv_rows2, now = transform_rows(self.raw_rows2)
            tickets = [r.get("Ticket No.", "").lstrip("0") or "0" for r in self.conv_rows2]
            self.gen_csv2 = self.gen_pdf2 = self.gen_mt_xlsx2 = self.gen_anesu_xlsx2 = None
            self.lbl_csv2_done.config(text=""); self.lbl_pdf2_done.config(text="")
            self.btn_csv2.config(text="Download CSV",bg=FBC_MID)
            self.btn_pdf2.config(text="Download PDF",bg=RED_DARK)
            fname=os.path.basename(path); exch2=get_exchange(self.raw_rows2[0].get("Market",""))
            for w in self.info_bar2.winfo_children(): w.pack_forget()
            self.info_bar2.pack(fill="x",pady=(0,6))
            tk.Label(self.info_bar2,text="OK",bg=TAG_BLUE,font=("Segoe UI",10)).pack(side="left",padx=(6,2),pady=4)
            self.lbl_file2.config(text=fname); self.lbl_file2.pack(side="left",pady=4)
            self.lbl_rows2.config(text=f"  {len(self.conv_rows2)} rows  |  {tickets[0]} to {tickets[-1]}")
            self.lbl_rows2.pack(side="left",pady=4); self.btn_reupload2.pack(side="right",padx=6,pady=4)
            for b in (self.btn_csv2,self.btn_pdf2): b.config(state="normal")
            self.btn_email2.config(text=f"Send — {exch2} Only",state="normal")
            self.prev_outer2.pack(fill="x",padx=12,pady=(4,0))
            self._build_preview("prev_body2","lbl_showing2",self.conv_rows2,tickets,now)
            if self.source_path: self.btn_email_both.config(state="normal")
        except Exception as e: messagebox.showerror("Error loading 2nd file",str(e))

    def _dl_csv2(self):
        try:
            exch2=get_exchange(self.raw_rows2[0].get("Market",""))
            self.gen_csv2=generate_csv(self.conv_rows2,self.out_dir,exch2)
            self.lbl_csv2_done.config(text=f"OK  {os.path.basename(self.gen_csv2)} saved")
            self.btn_csv2.config(text="CSV Downloaded",bg="#1B5E20")
        except Exception as e: messagebox.showerror("CSV Error (2nd)",str(e))

    def _dl_pdf2(self):
        try:
            self.gen_pdf2=generate_pdf(self.raw_rows2,self.raw_headers2,self.out_dir)
            self.lbl_pdf2_done.config(text=f"OK  {os.path.basename(self.gen_pdf2)} saved")
            self.btn_pdf2.config(text="PDF Downloaded",bg="#7B1010")
        except Exception as e: messagebox.showerror("PDF Error (2nd)",str(e))

    def _ensure_email_files2(self):
        if not self.gen_pdf2:
            self.gen_pdf2 = generate_pdf(self.raw_rows2, self.raw_headers2, self.out_dir)
            self.lbl_pdf2_done.config(text=f"OK  {os.path.basename(self.gen_pdf2)} saved")
        if not self.gen_mt_xlsx2:
            self.gen_mt_xlsx2 = generate_matched_excel(self.source_path2, self.raw_rows2, self.out_dir)
        if not self.gen_anesu_xlsx2:
            self.gen_anesu_xlsx2 = generate_anesu_excel(self.raw_rows2, self.out_dir)

    def _send_email2(self):
        try:
            self._ensure_email_files2()
            open_sarestock_outlook([self.gen_pdf2, self.gen_mt_xlsx2, self.gen_anesu_xlsx2],
                                   sender_name=load_sender_name())
        except ImportError: messagebox.showerror("pywin32 not installed","Run:  pip install pywin32")
        except Exception as e: messagebox.showerror("Outlook Error",str(e))

    def _send_email_both(self):
        try:
            self._ensure_email_files();
            self._ensure_email_files2()
            open_sarestock_outlook([self.gen_pdf, self.gen_mt_xlsx, self.gen_anesu_xlsx,
                                    self.gen_pdf2, self.gen_mt_xlsx2, self.gen_anesu_xlsx2],
                                   sender_name=load_sender_name())
        except ImportError: messagebox.showerror("pywin32 not installed","Run:  pip install pywin32")
        except Exception as e: messagebox.showerror("Outlook Error",str(e))

    def _pick_outdir(self):
        d=filedialog.askdirectory(title="Choose output folder",initialdir=self.out_dir)
        if d: self.out_dir=d; self.lbl_outdir.config(text=d)
# ════════════════════════════════════════════════════════════════════════════
#  EMAILER PAGE
# ════════════════════════════════════════════════════════════════════════════
class EmailerPage(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=BG)
        self.contacts=load_contacts(); self.deal_items=[]; self.pdf_folder=""
        self.sender_name = load_sender_name()
        self.sent_custodians = set()
        self.sent_clients    = set()
        self._build()

    def _build(self):
        bar=tk.Frame(self,bg=FBC_MID,padx=16,pady=8); bar.pack(fill="x")
        tk.Label(bar,text="Deal Note Email Automator",bg=FBC_MID,fg=WHITE,
                 font=("Segoe UI",11,"bold")).pack(side="left")
        tk.Button(bar,text="Manage Client Contacts",command=self._open_contacts,
                  bg=FBC_DARK,fg=WHITE,relief="flat",font=("Segoe UI",9,"bold"),
                  cursor="hand2",padx=10,pady=4).pack(side="right",padx=4)

        name_bar = tk.Frame(self, bg=FBC_DARK, padx=16, pady=6)
        name_bar.pack(fill="x")
        tk.Label(name_bar, text="Your Name (used in email sign-off):",
                 bg=FBC_DARK, fg=SIDEBAR_TEXT, font=("Segoe UI", 9, "bold")).pack(side="left")
        self._sender_var = tk.StringVar(value=self.sender_name)
        name_entry = tk.Entry(name_bar, textvariable=self._sender_var,
                              font=("Segoe UI", 10), bg="#0D2B4E", fg=WHITE,
                              insertbackground=WHITE, relief="flat",
                              highlightbackground=FBC_MID, highlightthickness=1,
                              width=26)
        name_entry.pack(side="left", padx=(8, 6), ipady=4)
        tk.Button(name_bar, text="Save Name", command=self._save_sender_name,
                  bg=GREEN_DARK, fg=WHITE, relief="flat",
                  font=("Segoe UI", 9, "bold"), cursor="hand2",
                  padx=10, pady=4).pack(side="left")
        self.lbl_name_saved = tk.Label(name_bar, text="", bg=FBC_DARK, fg="#90EE90",
                                       font=("Segoe UI", 8))
        self.lbl_name_saved.pack(side="left", padx=8)
        self.lbl_name_hint = tk.Label(name_bar,
            text=f"Saved: {self.sender_name}" if self.sender_name else "No name saved yet",
            bg=FBC_DARK,
            fg="#90CAF9" if self.sender_name else "#FF9966",
            font=("Segoe UI", 8))
        self.lbl_name_hint.pack(side="right")

        fp=tk.Frame(self,bg=WHITE,padx=16,pady=12); fp.pack(fill="x",padx=16,pady=(12,0))
        btn_row=tk.Frame(fp,bg=WHITE); btn_row.pack(fill="x")
        tk.Button(btn_row,text="Choose Deal Notes Folder",command=self._pick_folder,
                  bg=FBC_MID,fg=WHITE,relief="flat",font=("Segoe UI",10,"bold"),
                  cursor="hand2",padx=14,pady=8).pack(side="left")
        tk.Label(btn_row,text="  or  ",bg=WHITE,fg="#8096B0",font=("Segoe UI",9)).pack(side="left")
        tk.Button(btn_row,text="Select Individual Deal Note(s)",command=self._pick_individual_files,
                  bg="#4051B5",fg=WHITE,relief="flat",font=("Segoe UI",10,"bold"),
                  cursor="hand2",padx=14,pady=8).pack(side="left")
        self.btn_clear=tk.Button(btn_row,text="Clear All Uploads",command=self._clear_uploads,
                  bg=RED_DARK,fg=WHITE,relief="flat",font=("Segoe UI",9,"bold"),
                  cursor="hand2",padx=10,pady=8,state="disabled")
        self.btn_clear.pack(side="right")
        info_row=tk.Frame(fp,bg=WHITE); info_row.pack(fill="x",pady=(6,0))
        self.lbl_folder=tk.Label(info_row,text="No files loaded",bg=WHITE,fg="#8096B0",font=("Segoe UI",9))
        self.lbl_folder.pack(side="left")
        self.lbl_found=tk.Label(info_row,text="",bg=WHITE,fg=FBC_MID,font=("Consolas",9))
        self.lbl_found.pack(side="left",padx=10)
        self.lbl_file_list=tk.Label(info_row,text="",bg=WHITE,fg="#607080",font=("Segoe UI",8))
        self.lbl_file_list.pack(side="left")

        style=ttk.Style()
        style.configure("TNotebook.Tab",font=("Segoe UI",10,"bold"),padding=[14,6])
        nb=ttk.Notebook(self); nb.pack(fill="both",expand=True,padx=16,pady=10)
        self.tab_cust=tk.Frame(nb,bg=BG)
        self.tab_client=tk.Frame(nb,bg=BG)
        nb.add(self.tab_cust,text="  Custodian Emails  ")
        nb.add(self.tab_client,text="  Client Emails  ")
        self._build_custodian_tab()
        self._build_client_tab()

    # ── Custodian tab ────────────────────────────────────────────────────────
    def _build_custodian_tab(self):
        p=self.tab_cust
        tk.Label(p,text="Groups all PDFs by custodian - one email per custodian with all their deal notes attached.",
                 bg=BG,fg="#607080",font=("Segoe UI",9)).pack(anchor="w",padx=16,pady=(10,0))
        self.btn_send_all_cust=tk.Button(p,text="Send ALL Custodian Emails",
            command=self._cust_send_all,bg=GREEN_DARK,fg=WHITE,font=("Segoe UI",11,"bold"),
            relief="flat",padx=16,pady=10,cursor="hand2",state="disabled")
        self.btn_send_all_cust.pack(fill="x",padx=16,pady=(8,4))
        self.lbl_cust_hint=tk.Label(p,text="Load a folder above to begin.",bg=BG,fg="#607080",font=("Segoe UI",9))
        self.lbl_cust_hint.pack(anchor="w",padx=16,pady=(0,8))
        canvas=tk.Canvas(p,bg=BG,highlightthickness=0)
        sb=ttk.Scrollbar(p,orient="vertical",command=canvas.yview)
        self.cust_body=tk.Frame(canvas,bg=BG)
        self.cust_body.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0),window=self.cust_body,anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left",fill="both",expand=True,padx=(16,0))
        sb.pack(side="right",fill="y",pady=4)

    def _render_custodian_tab(self):
        for w in self.cust_body.winfo_children(): w.destroy()
        self.cust_btn_map={}
        groups={}
        for item in self.deal_items: groups.setdefault(item["custodian"],[]).append(item)
        for code,items in sorted(groups.items()):
            routing = get_effective_custodian_routing(code)
            sent = code in self.sent_custodians
            head_color = "#1A6B3A" if sent else (FBC_MID if routing else RED_DARK)
            card_bg    = "#F0FBF4" if sent else WHITE
            card=tk.Frame(self.cust_body,bg=card_bg,pady=0,padx=0)
            card.pack(fill="x",padx=4,pady=(0,10))
            head=tk.Frame(card,bg=head_color,pady=7,padx=12); head.pack(fill="x")
            label=routing["label"] if routing else "UNKNOWN CUSTODIAN"
            count=len(items)
            status_badge = "  ✓ SENT" if sent else ""
            tk.Label(head,text=f"{code}  —  {label}{status_badge}",bg=head_color,fg=WHITE,
                     font=("Segoe UI",10,"bold")).pack(side="left")
            tk.Label(head,text=f"{count} deal note{'s' if count>1 else ''}",
                     bg=head_color,fg=WHITE,font=("Segoe UI",9)).pack(side="right")
            inner=tk.Frame(card,bg=card_bg,padx=12,pady=8); inner.pack(fill="x")
            for it in items:
                tk.Label(inner,text=f"  {it['fname']}",bg=card_bg,fg="#2D3748",font=("Segoe UI",9)).pack(anchor="w")
            if routing:
                subj=f"DEAL NOTE{'S' if count>1 else ''} - {datetime.now().strftime('%d %b %Y')}"
                tk.Label(inner,text=f"Subject: {subj}",bg=card_bg,fg="#607080",
                         font=("Segoe UI",8,"italic")).pack(anchor="w",pady=(6,0))
                tk.Label(inner,text=f"To: {'; '.join(routing['to'])}",bg=card_bg,
                         fg="#607080",font=("Segoe UI",8)).pack(anchor="w")
                cc_summary = f"CC: {len(routing['cc'])} address{'es' if len(routing['cc'])!=1 else ''}"
                tk.Label(inner,text=cc_summary,bg=card_bg,fg="#8096B0",font=("Segoe UI",8)).pack(anchor="w")

                action_row = tk.Frame(inner, bg=card_bg)
                action_row.pack(anchor="w", pady=(8,0))

                # Send button
                btn_text = f"✓ Sent  ({count} file{'s' if count>1 else ''})" if sent \
                           else f"Open in Outlook  ({count} file{'s' if count>1 else ''})"
                btn_bg   = "#2E7D32" if sent else FBC_MID
                btn=tk.Button(action_row, text=btn_text,
                    command=lambda c=code: self._cust_send_one(c),
                    bg=btn_bg, fg=WHITE, relief="flat",
                    font=("Segoe UI",9,"bold"),
                    cursor="arrow" if sent else "hand2",
                    state="disabled" if sent else "normal",
                    disabledforeground=WHITE,
                    padx=10, pady=6)
                btn.pack(side="left", padx=(0, 8))
                self.cust_btn_map[code] = btn

                # Edit Recipients button (always visible per custodian)
                tk.Button(action_row, text="⚙ Edit Recipients",
                          command=lambda c=code: self._edit_custodian_recipients(c),
                          bg=FBC_ACCENT, fg=WHITE, relief="flat",
                          font=("Segoe UI",8,"bold"), cursor="hand2",
                          padx=8, pady=6).pack(side="left")
            else:
                tk.Label(inner,text="No routing configured for this custodian code.",
                         bg=card_bg,fg=RED_DARK,font=("Segoe UI",9)).pack(anchor="w")

        known  = sum(1 for c in groups if c in CUSTODIAN_ROUTING)
        unsent = sum(1 for c in groups if c in CUSTODIAN_ROUTING and c not in self.sent_custodians)
        self.btn_send_all_cust.config(
            state="normal" if unsent else "disabled",
            text=f"Send ALL {known} Custodian Email{'s' if known!=1 else ''} in Outlook"
                 + (f"  ({known - unsent} already sent)" if known > unsent else ""))
        self.lbl_cust_hint.config(
            text=f"  {len(self.deal_items)} deal note(s) across {len(groups)} custodian(s).",
            fg=GREEN_DARK)

    def _edit_custodian_recipients(self, code):
        """Open RecipientsDialog pre-filled with that custodian's current To/CC."""
        routing = get_effective_custodian_routing(code)
        if not routing:
            messagebox.showwarning("Unknown", f"No default routing for {code}.", parent=self)
            return
        def on_save(new_to, new_cc):
            overrides = load_custodian_overrides()
            overrides[code] = {"to": new_to or routing["to"], "cc": new_cc}
            save_custodian_overrides(overrides)
            self._render_custodian_tab()   # refresh cards to show new addresses
            messagebox.showinfo("Saved",
                f"Recipients for {code} updated.\nNew addresses will be used on the next send.",
                parent=self)
        RecipientsDialog(
            self,
            title=f"⚙  {routing['label']} — Edit Recipients",
            to_list=routing["to"],
            cc_list=routing["cc"],
            on_save=on_save,
            to_label=f"To  (primary recipients for {code})",
            cc_label=f"CC  (copied recipients for {code})",
        )

    def _cust_send_one(self,code):
        routing = get_effective_custodian_routing(code)
        if not routing: messagebox.showwarning("Unknown",f"No routing for {code}."); return
        items=[it for it in self.deal_items if it["custodian"]==code]
        count=len(items)
        subj=f"DEAL NOTE{'S' if count>1 else ''} - {datetime.now().strftime('%d %b %Y')}"
        body=get_custodian_body(multi=(count>1), sender_name=self._get_sender())
        try:
            open_outlook(routing["to"],routing["cc"],subj,body,[it["path"] for it in items])
            self.sent_custodians.add(code)
            if code in self.cust_btn_map:
                self.cust_btn_map[code].config(text="✓ Sent",bg="#2E7D32")
                self.after(300, self._render_custodian_tab)
        except ImportError as e: messagebox.showerror("pywin32 not installed",str(e))
        except Exception as e: messagebox.showerror("Outlook Error",str(e))

    def _cust_send_all(self):
        codes=sorted(set(it["custodian"] for it in self.deal_items
                         if it["custodian"] in CUSTODIAN_ROUTING
                         and it["custodian"] not in self.sent_custodians))
        if not codes: messagebox.showinfo("Nothing","No unsent custodian emails found."); return
        if not messagebox.askyesno("Send All",f"Open {len(codes)} Outlook window(s), one per custodian?\n\nContinue?"): return
        for code in codes: self._cust_send_one(code)
        messagebox.showinfo("Done",f"  {len(codes)} Outlook window(s) opened.")

    # ── Client tab ───────────────────────────────────────────────────────────
    def _build_client_tab(self):
        p=self.tab_client

        # ── Configure Client CC button row ────────────────────────────────
        cc_bar = tk.Frame(p, bg=BG)
        cc_bar.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(cc_bar,
                 text="Sends one email per client with all their deal notes attached. "
                      "Add emails via 'Manage Client Contacts'.",
                 bg=BG, fg="#607080", font=("Segoe UI",9)).pack(side="left")
        tk.Button(cc_bar, text="⚙ Configure CC",
                  command=self._configure_client_cc,
                  bg=FBC_ACCENT, fg=WHITE, relief="flat",
                  font=("Segoe UI", 8, "bold"), cursor="hand2",
                  padx=8, pady=4).pack(side="right")

        self.btn_send_everything=tk.Button(p,text="Send ALL Emails (Custodian + Client)",
            command=self._send_everything,bg=FBC_DARK,fg=WHITE,font=("Segoe UI",11,"bold"),
            relief="flat",padx=16,pady=10,cursor="hand2",state="disabled")
        self.btn_send_everything.pack(fill="x",padx=16,pady=(8,2))
        self.btn_send_all_client=tk.Button(p,text="Send ALL Client Emails",
            command=self._client_send_all,bg=GREEN_DARK,fg=WHITE,font=("Segoe UI",11,"bold"),
            relief="flat",padx=16,pady=10,cursor="hand2",state="disabled")
        self.btn_send_all_client.pack(fill="x",padx=16,pady=(2,4))
        self.lbl_client_hint=tk.Label(p,text="Load a folder above to begin.",bg=BG,fg="#607080",font=("Segoe UI",9))
        self.lbl_client_hint.pack(anchor="w",padx=16,pady=(0,4))
        wrap=tk.Frame(p,bg=BG); wrap.pack(fill="both",expand=True,padx=16,pady=4)
        canvas=tk.Canvas(wrap,bg=BG,highlightthickness=0)
        sb=ttk.Scrollbar(wrap,orient="vertical",command=canvas.yview)
        self.client_body=tk.Frame(canvas,bg=BG)
        self.client_body.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0),window=self.client_body,anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left",fill="both",expand=True); sb.pack(side="right",fill="y")

    def _configure_client_cc(self):
        """Open RecipientsDialog (CC only) for the client email CC list."""
        cc_list = load_client_cc()
        def on_save(_to, new_cc):
            save_client_cc(new_cc)
            messagebox.showinfo("Saved",
                f"Client email CC updated — {len(new_cc)} address(es) saved.\n\n"
                "All future client emails will use the new CC list.", parent=self)
        RecipientsDialog(
            self,
            title="⚙  Client Emails — Configure CC",
            to_list=None,          # CC-only mode
            cc_list=cc_list,
            on_save=on_save,
            cc_label="CC  (all client emails will copy these addresses)",
        )

    def _client_groups(self):
        groups = {}
        for item in self.deal_items:
            client_raw = item["client"]

            # ── Find canonical key: check contacts first ──────────────────────
            key = None
            for saved in self.contacts:
                if _names_match(client_raw, saved):
                    key = saved
                    break
            if key is None:
                for saved in self.contacts:
                    if client_raw.startswith(saved) or saved.startswith(client_raw):
                        key = saved
                        break

            # ── If not in contacts, normalize by sorting tokens so
            #    "EDWARD MTEWEYI" and "MTEWEYI EDWARD" map to the same key ─────
            if key is None:
                # Check if an existing group key matches by token set
                for existing_key in groups:
                    if _names_match(client_raw, existing_key):
                        key = existing_key
                        break

            if key is None:
                key = client_raw  # new unique client

            contact = find_contact(self.contacts, key)
            email = contact.get("email", "")

            if key not in groups:
                groups[key] = {"email": email, "items": [], "sent": False}
            groups[key]["items"].append(item)
            groups[key]["email"] = groups[key]["email"] or email

        for g in groups.values():
            g["status"] = "ready" if g["email"] else "missing"
        return groups

    def _render_client_tab(self):
        for w in self.client_body.winfo_children(): w.destroy()
        self.client_group_btns={}
        if not self.deal_items: return
        groups=self._client_groups()
        headers=["#","Client Name","Files","Custodian","Client Email","Status","Action"]
        widths=[3,22,18,10,26,9,10]
        hdr=tk.Frame(self.client_body,bg=FBC_DARK); hdr.pack(fill="x")
        for h,w in zip(headers,widths):
            tk.Label(hdr,text=h,bg=FBC_DARK,fg=WHITE,font=("Segoe UI",8,"bold"),
                     width=w,anchor="w",padx=4,pady=6).pack(side="left")
        for i,(client_name,grp) in enumerate(groups.items()):
            email=grp["email"]; status=grp["status"]; count=len(grp["items"])
            custodians=", ".join(sorted(set(it["custodian"] for it in grp["items"])))
            sent = client_name in self.sent_clients
            bg = "#E8F8EE" if sent else ("#F8FBFF" if i%2==0 else WHITE)
            row=tk.Frame(self.client_body,bg=bg); row.pack(fill="x")
            sc = "#2E7D32" if sent else (FBC_MID if status=="ready" else RED_DARK)
            st = "Sent" if sent else ("Pending" if status=="ready" else "Missing")
            file_label=f"{count} file{'s' if count>1 else ''}"
            for v,w in zip([str(i+1),client_name[:20],file_label,custodians[:10],email[:24] or "-",st],widths):
                tk.Label(row,text=v,bg=bg,fg=(sc if v==st else "#2D3748"),
                         font=("Segoe UI",8),width=w,anchor="w",padx=4,pady=5).pack(side="left")
            if sent:
                btn=tk.Button(row,text="✓ Sent",bg="#2E7D32",fg=WHITE,relief="flat",
                    font=("Segoe UI",8,"bold"),cursor="arrow",padx=6,pady=3,
                    state="disabled",disabledforeground=WHITE)
                btn.config(bg="#2E7D32")
            else:
                btn=tk.Button(row,text="Send",bg=FBC_MID,fg=WHITE,relief="flat",
                    font=("Segoe UI",8,"bold"),cursor="hand2",padx=6,pady=3,
                    state="normal" if status=="ready" else "disabled",
                    disabledforeground="#8096B0",
                    command=lambda cn=client_name:self._client_send_group(cn))
            btn.pack(side="left",padx=4); self.client_group_btns[client_name]=btn
        ready=sum(1 for g in groups.values() if g["status"]=="ready")
        unsent_ready=sum(1 for cn,g in groups.items() if g["status"]=="ready" and cn not in self.sent_clients)
        missing=len(groups)-ready
        self.btn_send_all_client.config(
            state="normal" if unsent_ready else "disabled",
            text=f"Send ALL {ready} Client Email{'s' if ready!=1 else ''} in Outlook"
                 + (f"  ({ready - unsent_ready} already sent)" if ready > unsent_ready else ""))
        self.btn_send_everything.config(
            state="normal" if ready or any(it["custodian"] in CUSTODIAN_ROUTING for it in self.deal_items) else "disabled")
        self.lbl_client_hint.config(
            text=(f"  All {ready} clients matched." if not missing
                  else f"  {missing} client(s) missing - click 'Manage Client Contacts'."),
            fg=GREEN_DARK if not missing else RED_DARK)

    def _client_send_group(self,client_name):
        groups=self._client_groups(); grp=groups.get(client_name)
        if not grp: return
        if grp["status"]=="missing":
            messagebox.showwarning("Missing Email",f"No email for '{client_name}'.\n\nUse 'Manage Client Contacts'."); return
        count=len(grp["items"])
        subj=f"DEAL{'S' if count>1 else ''} CONFIRMATION"
        body=get_client_body(client=client_name.title(), multi=(count>1), sender_name=self._get_sender())
        paths=[it["path"] for it in grp["items"]]
        client_cc = load_client_cc()   # always use saved CC
        try:
            open_outlook([grp["email"]], client_cc, subj, body, paths)
            self.sent_clients.add(client_name)
            if client_name in self.client_group_btns:
                self.client_group_btns[client_name].config(text="✓ Sent",bg="#2E7D32")
            self.after(300, self._render_client_tab)
        except ImportError as e: messagebox.showerror("pywin32 not installed",str(e))
        except Exception as e: messagebox.showerror("Outlook Error",str(e))

    def _client_send_all(self):
        groups=self._client_groups()
        ready=[cn for cn,g in groups.items()
               if g["status"]=="ready" and cn not in self.sent_clients]
        if not ready: messagebox.showinfo("Nothing","No unsent clients with emails found."); return
        if not messagebox.askyesno("Send All Client Emails",
                f"Open {len(ready)} Outlook window(s), one per client?\n\nContinue?"): return
        for cn in ready: self._client_send_group(cn)
        messagebox.showinfo("Done",f"  {len(ready)} Outlook window(s) opened.")

    def _send_everything(self):
        cust_codes=sorted(set(it["custodian"] for it in self.deal_items if it["custodian"] in CUSTODIAN_ROUTING))
        groups=self._client_groups(); ready_clients=[cn for cn,g in groups.items() if g["status"]=="ready"]
        total=len(cust_codes)+len(ready_clients)
        if total==0: messagebox.showinfo("Nothing","No emails to send."); return
        if not messagebox.askyesno("Send ALL Emails",
            f"This will open {total} Outlook window(s):\n"
            f"  - {len(cust_codes)} custodian email(s)\n"
            f"  - {len(ready_clients)} client email(s)\n\nContinue?"): return
        for code in cust_codes: self._cust_send_one(code)
        for cn in ready_clients: self._client_send_group(cn)
        messagebox.showinfo("Done",f"  {total} Outlook window(s) opened.")

    def _open_contacts(self):
        ContactsDialog(self,self.contacts,self._on_contacts_saved)

    def _on_contacts_saved(self,new_contacts):
        self.contacts=new_contacts
        if self.deal_items: self._render_client_tab()

    def _save_sender_name(self):
        name = self._sender_var.get().strip()
        if not name:
            messagebox.showwarning("Empty Name","Please enter your name before saving.", parent=self)
            return
        self.sender_name = name
        save_sender_name(name)
        self.lbl_name_saved.config(text="Saved!")
        self.lbl_name_hint.config(text=f"Saved: {name}", fg="#90CAF9")
        self.after(2500, lambda: self.lbl_name_saved.config(text=""))

    def _get_sender(self):
        return self._sender_var.get().strip() or self.sender_name or "FBC Securities"

    def _pick_folder(self):
        folder=filedialog.askdirectory(title="Select folder containing deal note PDFs")
        if not folder: return
        pdfs=sorted(f for f in os.listdir(folder) if f.lower().endswith(".pdf"))
        if not pdfs: messagebox.showwarning("No PDFs","No PDF files found in that folder."); return
        self.pdf_folder=folder
        self.lbl_folder.config(text=f"  {os.path.basename(folder)}",fg=FBC_DARK)
        self.lbl_found.config(text=f"Scanning {len(pdfs)} PDF(s)...")
        self.lbl_file_list.config(text="")
        self._disable_send_buttons()
        self.btn_clear.config(state="normal")
        threading.Thread(target=self._scan,args=(folder,pdfs),daemon=True).start()

    def _pick_individual_files(self):
        paths=filedialog.askopenfilenames(
            title="Select Deal Note PDF(s)",
            filetypes=[("PDF files","*.pdf"),("All files","*.*")])
        if not paths: return
        pdf_paths=list(paths)
        already={it["path"] for it in self.deal_items}
        new_paths=[p for p in pdf_paths if p not in already]
        if not new_paths:
            messagebox.showinfo("No New Files","All selected files are already loaded."); return
        self.lbl_found.config(text=f"Scanning {len(new_paths)} new PDF(s)...")
        self.btn_clear.config(state="normal")
        names=", ".join(os.path.basename(p) for p in new_paths[:3])
        if len(new_paths)>3: names+=f" +{len(new_paths)-3} more"
        self.lbl_folder.config(text=f"  {names}",fg=FBC_DARK)
        threading.Thread(target=self._scan_files,args=(new_paths,),daemon=True).start()

    def _clear_uploads(self):
        if not self.deal_items: return
        if not messagebox.askyesno("Clear All Uploads",
            f"Remove all {len(self.deal_items)} loaded deal note(s) and start fresh?\n\n"
            "This does NOT delete the files from disk."): return
        self.deal_items=[]; self.pdf_folder=""
        self.sent_custodians.clear(); self.sent_clients.clear()
        self.lbl_folder.config(text="No files loaded",fg="#8096B0")
        self.lbl_found.config(text=""); self.lbl_file_list.config(text="")
        self.btn_clear.config(state="disabled")
        self._disable_send_buttons()
        for w in self.cust_body.winfo_children(): w.destroy()
        for w in self.client_body.winfo_children(): w.destroy()
        self.lbl_cust_hint.config(text="Load files above to begin.",fg="#607080")
        self.lbl_client_hint.config(text="Load files above to begin.",fg="#607080")
        self.btn_send_all_cust.config(state="disabled",text="Send ALL Custodian Emails")
        self.btn_send_all_client.config(state="disabled",text="Send ALL Client Emails")
        self.btn_send_everything.config(state="disabled")

    def _disable_send_buttons(self):
        self.btn_send_all_cust.config(state="disabled")
        self.btn_send_all_client.config(state="disabled")
        self.btn_send_everything.config(state="disabled")

    def _scan(self,folder,pdfs):
        items=[]
        for fname in pdfs:
            path=os.path.join(folder,fname)
            client_from_pdf = parse_client_name_from_pdf(path)
            client = client_from_pdf if client_from_pdf else parse_client_name_from_filename(fname)
            items.append({
                "fname":fname,"path":path,"client":client,
                "client_source": "pdf" if client_from_pdf else "filename",
                "custodian":parse_custodian_from_pdf(path) or "UNKNOWN",
                "deal_info":parse_deal_info_from_pdf(path),"sent":False,
            })
        self.deal_items=items
        self._after_scan(len(items))

    def _scan_files(self,paths):
        new_items=[]
        for path in paths:
            fname=os.path.basename(path)
            client_from_pdf = parse_client_name_from_pdf(path)
            client = client_from_pdf if client_from_pdf else parse_client_name_from_filename(fname)
            new_items.append({
                "fname":fname,"path":path,"client":client,
                "client_source": "pdf" if client_from_pdf else "filename",
                "custodian":parse_custodian_from_pdf(path) or "UNKNOWN",
                "deal_info":parse_deal_info_from_pdf(path),"sent":False,
            })
        self.deal_items.extend(new_items)
        self._after_scan(len(self.deal_items))

    def _after_scan(self,total):
        self.after(0,self._render_custodian_tab)
        self.after(0,self._render_client_tab)
        self.after(0,lambda:self.lbl_found.config(text=f"  {total} PDF(s) loaded"))


# ════════════════════════════════════════════════════════════════════════════
#  VOICE ENGINE
# ════════════════════════════════════════════════════════════════════════════
_tts_lock = threading.Lock()

def speak(text: str):
    if not _VOICE_READY or _tts is None:
        return
    def _run():
        with _tts_lock:
            try:
                _tts.say(text); _tts.runAndWait()
            except Exception:
                pass
    threading.Thread(target=_run, daemon=True).start()


class VoiceBar(tk.Frame):
    MIC_IDLE   = ("🎤  Hold to speak", FBC_MID,   WHITE)
    MIC_LISTEN = ("🔴  Listening…",   "#B71C1C",  WHITE)
    MIC_THINK  = ("⏳  Processing…",  "#555555",  WHITE)
    MIC_NODEPS = ("🎤  Voice (install deps)", "#2A4A6A", SIDEBAR_TEXT)

    def __init__(self, parent, dispatch_cb, hotkey_widget=None):
        super().__init__(parent, bg=SIDEBAR_BG)
        self._cb     = dispatch_cb
        self._active = False
        self._build()
        if hotkey_widget:
            hotkey_widget.bind_all("<Control-space>", lambda e: self._toggle())

    def _build(self):
        tk.Frame(self, bg=FBC_MID, height=1).pack(fill="x", padx=10, pady=(8,4))
        tk.Label(self, text="VOICE ASSISTANT", bg=SIDEBAR_BG,
                 fg="#2A4A6A", font=("Segoe UI",7,"bold")).pack()
        self.btn = tk.Button(
            self,
            text=self.MIC_IDLE[0]   if _VOICE_READY else self.MIC_NODEPS[0],
            bg=self.MIC_IDLE[1]     if _VOICE_READY else self.MIC_NODEPS[1],
            fg=self.MIC_IDLE[2]     if _VOICE_READY else self.MIC_NODEPS[2],
            relief="flat", font=("Segoe UI",9,"bold"),
            cursor="hand2" if _VOICE_READY else "arrow",
            padx=6, pady=10, wraplength=170,
            command=self._toggle if _VOICE_READY else self._show_install,
        )
        self.btn.pack(fill="x", padx=10, pady=4)
        self.lbl_heard = tk.Label(self, text="", bg=SIDEBAR_BG, fg="#90CAF9",
                                  font=("Segoe UI",8), wraplength=170, justify="left")
        self.lbl_heard.pack(fill="x", padx=10)
        tk.Label(self, text="Ctrl+Space to activate",
                 bg=SIDEBAR_BG, fg="#2A4A6A", font=("Segoe UI",7)).pack(pady=(2,6))

    def _set_state(self, s):
        self.btn.config(text=s[0], bg=s[1], fg=s[2])

    def _show_install(self):
        messagebox.showinfo("Voice — Install Required",
            "To enable voice control, open a terminal and run:\n\n"
            "  pip install SpeechRecognition pyttsx3 pyaudio\n\n"
            "If pyaudio fails:\n"
            "  pip install pipwin\n"
            "  pipwin install pyaudio\n\n"
            "Then restart FBC Suite.")

    def _toggle(self):
        if self._active: return
        self._active = True
        self._set_state(self.MIC_LISTEN)
        threading.Thread(target=self._listen, daemon=True).start()

    def _listen(self):
        r = _sr.Recognizer(); mic = _sr.Microphone(); text = ""
        try:
            with mic as source:
                r.adjust_for_ambient_noise(source, duration=0.3)
                audio = r.listen(source, timeout=6, phrase_time_limit=8)
            self.after(0, lambda: self._set_state(self.MIC_THINK))
            text = r.recognize_google(audio).lower().strip()
        except _sr.WaitTimeoutError: text = ""
        except _sr.UnknownValueError: text = ""
        except Exception as exc:
            text = ""
            self.after(0, lambda: self.lbl_heard.config(text=f"Error: {exc}"))
        finally:
            self._active = False
            self.after(0, lambda: self._set_state(self.MIC_IDLE))
        if text:
            self.after(0, lambda t=text: self._on_heard(t))

    def _on_heard(self, text: str):
        self.lbl_heard.config(text=f'"{text}"')
        self._cb(text)
        self.after(6000, lambda: self.lbl_heard.config(text=""))


# ════════════════════════════════════════════════════════════════════════════
#  MAIN APP SHELL
# ════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"FBC Suite  v{VERSION}")
        self.state("zoomed")
        self.configure(bg=SIDEBAR_BG)
        self._active_page = None
        self._build()

    def _build(self):
        sidebar = tk.Frame(self, bg=SIDEBAR_BG, width=210)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        logo = tk.Frame(sidebar, bg=SIDEBAR_BG, pady=20)
        logo.pack(fill="x")
        tk.Label(logo, text="FBC", bg=FBC_ACCENT, fg=WHITE,
                 font=("Segoe UI",16,"bold"), padx=10, pady=6).pack()
        tk.Label(logo, text="Suite", bg=SIDEBAR_BG, fg=SIDEBAR_TEXT,
                 font=("Segoe UI",10)).pack(pady=(4,0))
        tk.Frame(sidebar, bg=FBC_MID, height=1).pack(fill="x", padx=16, pady=(0,10))
        self.nav_buttons = {}
        for label, key in [("Converter","converter"),("Deal Note\nEmailer","emailer")]:
            btn = tk.Button(sidebar, text=label,
                command=lambda k=key: self._switch(k),
                bg=SIDEBAR_BG, fg=SIDEBAR_TEXT,
                activebackground=SIDEBAR_ACTIVE, activeforeground=WHITE,
                font=("Segoe UI",10,"bold"), relief="flat",
                cursor="hand2", pady=16, width=16, justify="center")
            btn.pack(fill="x", padx=8, pady=2)
            btn.bind("<Enter>", lambda e,b=btn,k=key: b.config(
                bg=SIDEBAR_ACTIVE if self._active_page==k else SIDEBAR_HOVER))
            btn.bind("<Leave>", lambda e,b=btn,k=key: b.config(
                bg=SIDEBAR_ACTIVE if self._active_page==k else SIDEBAR_BG))
            self.nav_buttons[key] = btn
        self.voice_bar = VoiceBar(sidebar, self._voice_dispatch, hotkey_widget=self)
        self.voice_bar.pack(side="bottom", fill="x")
        tk.Label(sidebar, text=f"v{VERSION}", bg=SIDEBAR_BG, fg="#2A4A6A",
                 font=("Segoe UI",8)).pack(side="bottom", pady=4)
        self.content = tk.Frame(self, bg=BG)
        self.content.pack(side="left", fill="both", expand=True)
        self.pages = {
            "converter": SarestockPage(self.content),
            "emailer":   EmailerPage(self.content),
        }
        self._switch("converter")

    def _voice_dispatch(self, text: str):
        t = text.lower()
        if any(w in t for w in ("converter","convert","sarestock","first tab")):
            self._switch("converter"); speak("Switched to Converter."); return
        if any(w in t for w in ("email","emailer","deal note","second tab")):
            self._switch("emailer"); speak("Switched to Deal Note Emailer."); return
        conv = self.pages["converter"]
        if any(w in t for w in ("browse","upload file","load file","open file","first exchange","pick file")):
            self._switch("converter"); speak("Opening file browser."); conv._pick_file(); return
        if any(w in t for w in ("second exchange","second file","upload second","load second","vfex file","pick second")):
            self._switch("converter"); speak("Opening second file browser."); conv._pick_file2(); return
        if any(w in t for w in ("download csv","save csv","get csv")):
            self._switch("converter")
            if conv.conv_rows: speak("Downloading CSV."); conv._dl_csv()
            elif conv.conv_rows2: speak("Downloading second CSV."); conv._dl_csv2()
            else: speak("No file loaded yet.")
            return
        if any(w in t for w in ("download pdf","save pdf","get pdf")):
            self._switch("converter")
            if conv.conv_rows: speak("Downloading PDF."); conv._dl_pdf()
            elif conv.conv_rows2: speak("Downloading second PDF."); conv._dl_pdf2()
            else: speak("No file loaded yet.")
            return
        if any(w in t for w in ("send zse","send z s e","zse email")):
            self._switch("converter"); speak("Opening ZSE email."); conv._send_email(); return
        if any(w in t for w in ("send vfex","vfex email")):
            self._switch("converter"); speak("Opening VFEX email."); conv._send_email2(); return
        if any(w in t for w in ("send both","both emails","zse and vfex","send everything converter")):
            self._switch("converter"); speak("Opening combined email."); conv._send_email_both(); return
        if any(w in t for w in ("clear converter","clear uploads converter")):
            self._switch("converter"); speak("Clearing converter uploads."); conv._clear_uploads(); return
        em = self.pages["emailer"]
        if any(w in t for w in ("load folder","pick folder","open folder","select folder","load pdfs","browse folder")):
            self._switch("emailer"); speak("Opening folder browser."); em._pick_folder(); return
        if any(w in t for w in ("load files","pick files","individual files","select files","browse files")):
            self._switch("emailer"); speak("Opening file picker."); em._pick_individual_files(); return
        if any(w in t for w in ("send all custodian","custodian emails","send custodians","all custodians")):
            self._switch("emailer"); speak("Sending all custodian emails."); em._cust_send_all(); return
        if any(w in t for w in ("send all client","client emails","send clients","all clients")):
            self._switch("emailer"); speak("Sending all client emails."); em._client_send_all(); return
        if any(w in t for w in ("send everything","send all emails","send all","everything")):
            self._switch("emailer"); speak("Sending everything."); em._send_everything(); return
        if any(w in t for w in ("manage contacts","contacts","open contacts")):
            self._switch("emailer"); speak("Opening contacts."); em._open_contacts(); return
        if any(w in t for w in ("clear emailer","clear deal","clear files emailer")):
            self._switch("emailer"); speak("Clearing loaded files."); em._clear_uploads(); return
        if any(w in t for w in ("help","commands","what can you do","what can i say")):
            speak("You can say: switch to converter, switch to emailer, browse file, download CSV, "
                  "download PDF, send ZSE, send VFEX, send both, load folder, send all custodians, "
                  "send all clients, send everything, manage contacts, or reset counter.")
            _VoiceHelpDialog(self); return
        if any(w in t for w in ("status","how many","how many files","what's loaded")):
            parts = []
            if conv.conv_rows: parts.append(f"{len(conv.conv_rows)} rows in first exchange")
            if conv.conv_rows2: parts.append(f"{len(conv.conv_rows2)} rows in second exchange")
            if em.deal_items: parts.append(f"{len(em.deal_items)} deal notes loaded")
            speak((", ".join(parts)+".") if parts else "Nothing loaded yet."); return
        speak("Sorry, I didn't catch that. Say 'help' for a list of commands.")

    def _switch(self, key):
        for page in self.pages.values(): page.pack_forget()
        self.pages[key].pack(fill="both", expand=True)
        self._active_page = key
        for k, btn in self.nav_buttons.items():
            btn.config(bg=SIDEBAR_ACTIVE if k==key else SIDEBAR_BG,
                       fg=WHITE         if k==key else SIDEBAR_TEXT)


class _VoiceHelpDialog(tk.Toplevel):
    COMMANDS = [
        ("Navigation",        ["switch to converter", "switch to emailer"]),
        ("Converter",         ["browse file / second file","download CSV / PDF",
                               "send ZSE / send VFEX / send both",
                               "reset counter","clear converter"]),
        ("Deal Note Emailer", ["load folder / load files","send all custodians",
                               "send all clients","send everything",
                               "manage contacts","clear emailer"]),
        ("General",           ["status — how many files loaded",
                               "help — show this dialog",
                               "Ctrl+Space — activate mic anywhere"]),
    ]
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Voice Commands — FBC Suite")
        self.configure(bg=SIDEBAR_BG)
        self.resizable(False, False)
        self.grab_set()
        tk.Label(self, text="🎤  Voice Commands", bg=FBC_ACCENT, fg=WHITE,
                 font=("Segoe UI",12,"bold"), pady=10).pack(fill="x")
        body = tk.Frame(self, bg=SIDEBAR_BG, padx=20, pady=14)
        body.pack(fill="both", expand=True)
        for section, cmds in self.COMMANDS:
            tk.Label(body, text=section.upper(), bg=SIDEBAR_BG, fg=FBC_ACCENT,
                     font=("Segoe UI",8,"bold")).pack(anchor="w", pady=(8,2))
            for c in cmds:
                tk.Label(body, text=f"  • {c}", bg=SIDEBAR_BG, fg=SIDEBAR_TEXT,
                         font=("Segoe UI",9)).pack(anchor="w")
        tk.Button(self, text="Close", command=self.destroy,
                  bg=FBC_MID, fg=WHITE, relief="flat",
                  font=("Segoe UI",10,"bold"), pady=8, cursor="hand2").pack(
                      fill="x", padx=20, pady=14)
        self.update_idletasks()
        w, h = 360, self.winfo_reqheight()
        x = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        y = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")


# ════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    check_and_apply_update()
    login = LoginDialog()
    login.mainloop()
    if not login.authenticated:
        sys.exit(0)
    app = App()
    app.mainloop()
